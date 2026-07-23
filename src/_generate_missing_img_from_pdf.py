"""
_generate_missing_img_from_pdf.py
为缺失产品图片、但 Excel 里有 NMPA Artwork PDF 链接（pdf_url）的产品，
下载 PDF 并渲染首页为 PNG，作为前端产品图的兜底显示。

用法: python src/_generate_missing_img_from_pdf.py [--limit N] [--dry-run]
"""
import sys, re, json, argparse
from pathlib import Path
import openpyxl
import fitz  # PyMuPDF

sys.path.insert(0, str(Path(__file__).parent))
from config import BRANDS, DOWNLOAD_BASE, COL_NAME, COL_PDF_URL, EXCEL_PATH
from module3_download import (
    download_file, resolve_nmpa_pdf_url, sanitize,
    HZPBA_HOST, _hzpba_download_with_page,
)

WORKSPACE = Path(__file__).parent.parent
OUT_DIR   = WORKSPACE / "res" / "product_images" / "_pdf_snapshot"
MAP_JSON  = WORKSPACE / "res" / "product_images" / "image_map.json"
OUT_DIR.mkdir(parents=True, exist_ok=True)

_REG_PAT = re.compile(r"(国妆特字|国妆特进字|卫妆特字|妆网备字|国妆备字|国妆备进字)")

# 已知的坏 OLE 图标截图（Word/Excel 里插入的是 PDF 对象图标，不是产品照片）—先清理
KNOWN_BAD = {
    "Kans": [
        "韩束清润净卸洁面膏", "韩束水光保湿弹润眼霜", "韩束闪充冰美式盈润醒肤喷雾",
        "韩束舒缓修护精华面膜", "韩束凝时多肽塑颜精华霜",
    ],
}


def excel_files():
    base_dir = Path(EXCEL_PATH).parent
    historical = sorted(base_dir.glob("CI_List_Ada *.xlsx"))
    current = Path(EXCEL_PATH)
    result = [f for f in historical if f != current]
    if current.exists():
        result.append(current)
    return result


def load_records():
    records = []
    seen = set()
    for path in excel_files():
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
        except Exception:
            continue
        for brand_en in BRANDS:
            if brand_en not in wb.sheetnames:
                continue
            ws = wb[brand_en]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 5:
                    continue

                def g(col):
                    return row[col - 1] if len(row) >= col else None

                name = g(COL_NAME)
                pdf_url = g(COL_PDF_URL)
                if not name or not pdf_url:
                    continue
                name = str(name).strip()
                pdf_url = str(pdf_url).strip()
                if not pdf_url or pdf_url in ("None", "NA"):
                    continue

                reg = ""
                for cell in row:
                    s = str(cell or "").strip()
                    if s and _REG_PAT.search(s):
                        reg = s
                        break
                key = reg or f"{brand_en}:{name}"
                if key in seen:
                    continue
                seen.add(key)
                is_special = bool(reg and re.search(r"国妆特字|国妆特进字|卫妆特字", reg))
                records.append({"brand_en": brand_en, "name": name,
                                 "pdf_url": pdf_url, "is_special": is_special})
        wb.close()
    return records


def load_image_map():
    if MAP_JSON.exists():
        return json.loads(MAP_JSON.read_text(encoding="utf-8"))
    return {}


def has_image(brand_en, name, img_map):
    brand_imgs = img_map.get(brand_en, {})
    if not brand_imgs:
        return False
    if name in brand_imgs:
        return True
    for key in brand_imgs:
        if len(key) >= 4 and (key in name or name in key):
            return True
    return False


def render_pdf_first_page(pdf_path: Path, out_png: Path, max_dim_px: int = 1000) -> bool:
    """渲染PDF首页为PNG，按物理页面最长边缩放到 max_dim_px 像素以内，
    避免某些PDF页面物理尺寸异常大导致生成的PNG体积过大（曾出现单张>100MB
    导致前端 base64 内嵌全部加载进内存时 MemoryError 的问题）。"""
    try:
        doc = fitz.open(str(pdf_path))
        if doc.page_count == 0:
            return False
        page = doc[0]
        rect = page.rect
        page_max = max(rect.width, rect.height, 1)
        scale = max_dim_px / page_max
        scale = max(min(scale, 3.0), 0.05)
        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale))
        pix.save(str(out_png))
        return True
    except Exception as e:
        print(f"  渲染失败 {pdf_path.name}: {e}")
        return False


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0, help="最多处理多少条（0=不限）")
    ap.add_argument("--dry-run", action="store_true", help="只统计，不下载不渲染")
    args = ap.parse_args()

    records = load_records()
    img_map = load_image_map()
    print(f"共 {len(records)} 条含 pdf_url 的唯一记录")

    # 先清理已知坏图标
    for brand_en, names in KNOWN_BAD.items():
        if brand_en in img_map:
            for nm in names:
                if img_map[brand_en].pop(nm, None):
                    print(f"已清理坏图标: {brand_en} / {nm}")

    todo = [r for r in records if not has_image(r["brand_en"], r["name"], img_map)]
    print(f"缺图片、待处理: {len(todo)} 条")

    if args.dry_run:
        by_brand = {}
        for r in todo:
            by_brand[r["brand_en"]] = by_brand.get(r["brand_en"], 0) + 1
        for b, n in sorted(by_brand.items(), key=lambda x: -x[1]):
            print(f"  {b:16s} {n}")
        MAP_JSON.write_text(json.dumps(img_map, ensure_ascii=False, indent=2), encoding="utf-8")
        return

    if args.limit:
        todo = todo[: args.limit]

    def paths_for(r):
        safe = sanitize(r["name"])
        prefix = "特化--" if r["is_special"] else ""
        pdf_path = DOWNLOAD_BASE / r["brand_en"] / f"{prefix}{safe}.pdf"
        out_dir = OUT_DIR / r["brand_en"]
        out_dir.mkdir(parents=True, exist_ok=True)
        return pdf_path, out_dir / f"{safe}.png"

    ok, fail = 0, 0

    def finalize(r, pdf_path, out_png):
        nonlocal ok, fail
        if render_pdf_first_page(pdf_path, out_png):
            img_map.setdefault(r["brand_en"], {})[r["name"]] = str(out_png).replace("\\", "/")
            ok += 1
        else:
            fail += 1

    hzpba_todo = [r for r in todo if HZPBA_HOST in r["pdf_url"]]
    nmpa_todo  = [r for r in todo if HZPBA_HOST not in r["pdf_url"]]

    def pdf_ready(p: Path) -> bool:
        return p.exists() and p.stat().st_size > 0

    def save_checkpoint():
        MAP_JSON.write_text(json.dumps(img_map, ensure_ascii=False, indent=2), encoding="utf-8")

    # ── nmpa datasearch（特殊注册）：直接 HTTP 下载，无需浏览器 ──
    for i, r in enumerate(nmpa_todo):
        pdf_path, out_png = paths_for(r)
        print(f"[nmpa {i+1}/{len(nmpa_todo)}] {r['brand_en']} / {r['name'][:24]}")
        if not pdf_ready(pdf_path):
            real_url = resolve_nmpa_pdf_url(r["pdf_url"])
            try:
                ok_dl = download_file(real_url or r["pdf_url"], pdf_path,
                                       referer="https://www.nmpa.gov.cn/")
            except Exception as e:
                print(f"  下载异常: {e}")
                ok_dl = False
            if not ok_dl:
                fail += 1
                continue
        finalize(r, pdf_path, out_png)
        if (ok + fail) % 10 == 0:
            save_checkpoint()

    # ── hzpba（普通备案）：单个 Playwright 浏览器实例，逐个复用同一 page ──
    if hzpba_todo:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True,
                args=["--disable-blink-features=AutomationControlled"],
            )
            ctx = browser.new_context(
                accept_downloads=True,
                locale="zh-CN",
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
            )
            page = ctx.new_page()
            try:
                page.goto("https://hzpba.nmpa.gov.cn/", timeout=20000)
                page.wait_for_timeout(1000)
            except Exception as e:
                print(f"hzpba 主页访问失败（继续尝试下载）: {e}")

            for i, r in enumerate(hzpba_todo):
                pdf_path, out_png = paths_for(r)
                print(f"[hzpba {i+1}/{len(hzpba_todo)}] {r['brand_en']} / {r['name'][:24]}")
                if not pdf_ready(pdf_path):
                    if not _hzpba_download_with_page(page, r["pdf_url"], pdf_path):
                        fail += 1
                        continue
                finalize(r, pdf_path, out_png)
                if (ok + fail) % 10 == 0:
                    save_checkpoint()

            browser.close()

    MAP_JSON.write_text(json.dumps(img_map, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n完成: 成功 {ok}，失败 {fail}")


if __name__ == "__main__":
    main()
