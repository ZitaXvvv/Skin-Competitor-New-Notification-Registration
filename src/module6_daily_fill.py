"""
module6_daily_fill.py
每日增量补全任务：检查连通性，按优先顺序（2026→2025→2024）每天补全 10 个
缺失信息的产品卡片（成分/PDF链接/产品图/mini-POC链接）；
并每周自动识别并补全完全缺失的历史月份数据（陆续补齐缺失月份）。

用法:
  python src/module6_daily_fill.py             # 正常运行（10个/天 + 每周月份补全）
  python src/module6_daily_fill.py --limit 5   # 自定义每次处理数量
  python src/module6_daily_fill.py --dry-run   # 只统计，不写入
  python src/module6_daily_fill.py --check     # 只做连通性检查后退出
  python src/module6_daily_fill.py --backfill  # 只运行历史月份补全（忽略每周间隔）
"""

import argparse
import io
import json
import logging
import re
import subprocess
import sys
import time
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from config import (
    BRANDS, EXCEL_PATH, DOWNLOAD_BASE,
    COL_UPLOAD_DATE, COL_NAME, COL_EFFECT, COL_DATE, COL_REG_NUM,
    COL_CATEGORY, COL_INGREDIENTS, COL_PDF_URL, COL_LABEL_URL, COL_POC_URL,
    HZPBA_SEARCH_URL, HZPBA_IMPORT_URL, HZPBA_PDF_BASE, LOG_DIR,
)

# ── 路径 ──────────────────────────────────────────────────────────────────
WORKSPACE        = Path(__file__).parent.parent
MAP_JSON         = WORKSPACE / "res" / "product_images" / "image_map.json"
OUT_IMG_DIR      = WORKSPACE / "res" / "product_images" / "_pdf_snapshot"
BACKFILL_STATE   = LOG_DIR / "backfill_state.json"   # 记录已尝试/完成的月份
MAIN_SCRIPT      = Path(__file__).parent / "main.py"
LOG_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(
            LOG_DIR / f"daily_fill_{datetime.today().strftime('%Y%m%d')}.log",
            encoding="utf-8",
        ),
    ],
)
log = logging.getLogger(__name__)

# 月份归档文件名模式：CI_List_Ada {月份缩写}'YY.xlsx
# 例：CI_List_Ada Jun'25.xlsx
_MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
# 从哪年开始追溯
BACKFILL_START_YEAR = 2024

# ── 缺失判断 ──────────────────────────────────────────────────────────────
_REG_PAT = re.compile(r"(妆网备字|国妆备字|国妆特字|国妆特进字|国妆备进字|卫妆特字)")
_EXCLUDE = re.compile(r"唇|口红|男士|男仕|契尔氏|名女人")
IMG_MAX_PX = 320
JPEG_Q     = 82


def _clean(v) -> str:
    s = str(v or "").strip()
    return "" if s in ("None", "NA") else s


def _is_special(reg: str) -> bool:
    return bool(re.search(r"国妆特字|国妆特进字|卫妆特字", reg))


def _is_imported(reg: str) -> bool:
    return "进" in reg


def _has_image(brand_en: str, name: str, img_map: dict) -> bool:
    brand_imgs = img_map.get(brand_en, {})
    if name in brand_imgs:
        return True
    short = name
    for zh in ["珀莱雅","谷雨","欧诗漫","兰蔻","欧莱雅","雅诗兰黛","修丽可",
               "百雀羚","韩束","自然堂","薇诺娜","科颜氏","娇韵诗"]:
        if name.startswith(zh):
            short = name[len(zh):]
            break
    for k in brand_imgs:
        ks = k
        for zh in ["珀莱雅","谷雨","欧诗漫","兰蔻","欧莱雅","雅诗兰黛","修丽可",
                   "百雀羚","韩束","自然堂","薇诺娜","科颜氏","娇韵诗"]:
            if k.startswith(zh):
                ks = k[len(zh):]
                break
        if len(ks) >= 4 and (ks in name or ks in short or short in ks):
            return True
    return False


def score_completeness(rec: dict, img_map: dict) -> int:
    """返回缺失字段数（0=完整）"""
    missing = 0
    if not rec.get("ingredients"):   missing += 1
    if not rec.get("pdf_url"):       missing += 1
    if not rec.get("poc_url"):       missing += 1
    if not _has_image(rec["brand_en"], rec["name"], img_map):
        missing += 1
    return missing


# ── 连通性检查 ────────────────────────────────────────────────────────────
def check_bebd(page) -> bool:
    """返回 True 表示 BEBD 已登录（session 有效）"""
    try:
        from config import BEBD_URL, COOKIES_FILE
        if COOKIES_FILE.exists():
            cookies = json.loads(COOKIES_FILE.read_text(encoding="utf-8"))
            page.context.add_cookies(cookies)
        page.goto(BEBD_URL, timeout=20000)
        page.wait_for_timeout(3000)
        content = page.content()
        logged_in = "退出" in content or "个人中心" in content or "我的" in content
        if logged_in:
            log.info("  BEBD: ✅ 已登录")
        else:
            log.warning("  BEBD: ⚠️ Cookie 已失效，需要手动重新登录")
        return logged_in
    except Exception as e:
        log.warning(f"  BEBD 连通性检查失败: {e}")
        return False


def check_nmpa(page) -> bool:
    """返回 True 表示 NMPA hzpba 可以正常访问"""
    try:
        page.goto(HZPBA_SEARCH_URL, timeout=20000)
        page.wait_for_timeout(3000)
        status = page.evaluate("() => document.readyState")
        ok = status == "complete" and "hzpba" in page.url
        if ok:
            log.info("  NMPA hzpba: ✅ 可访问")
        else:
            log.warning("  NMPA hzpba: ⚠️ 访问异常（WAF拦截或网络问题）")
        return ok
    except Exception as e:
        log.warning(f"  NMPA hzpba 连通性检查失败: {e}")
        return False


# ── 加载所有历史 + 当前 Excel ─────────────────────────────────────────────
def _all_excel_files() -> list:
    base = Path(EXCEL_PATH).parent
    historical = sorted(base.glob("CI_List_Ada *.xlsx"))
    current = Path(EXCEL_PATH)
    return [f for f in historical if f != current] + ([current] if current.exists() else [])


def load_incomplete_records(img_map: dict) -> list:
    """
    加载所有 Excel 文件中不完整的产品记录，
    按年份降序（2026→2025→2024）排序，去重后返回。
    """
    records = []
    seen_regs: set = set()

    for xlsx_path in _all_excel_files():
        try:
            wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        except Exception:
            continue

        for brand_en, brand_cn in BRANDS.items():
            if brand_en not in wb.sheetnames:
                continue
            ws = wb[brand_en]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 5:
                    continue

                def g(col):
                    return row[col - 1] if len(row) >= col else None

                name     = _clean(g(COL_NAME))
                if not name or _EXCLUDE.search(name):
                    continue

                reg_raw = ""
                for cell in row:
                    s = _clean(cell)
                    if s and _REG_PAT.search(s):
                        reg_raw = s
                        break
                if not reg_raw:
                    continue
                dedup_key = reg_raw
                if dedup_key in seen_regs:
                    continue
                seen_regs.add(dedup_key)

                # Parse year from notification date
                notif_raw  = g(COL_DATE)
                upload_raw = g(COL_UPLOAD_DATE)
                year = 0
                for val in [notif_raw, upload_raw]:
                    if val is None:
                        continue
                    s = str(val).strip()
                    m = re.search(r"\b(20\d{2})\b", s)
                    if m:
                        year = int(m.group(1))
                        break

                ingr   = _clean(g(COL_INGREDIENTS))
                pdf    = _clean(g(COL_PDF_URL))
                poc    = _clean(g(COL_POC_URL))
                label  = _clean(g(COL_LABEL_URL))

                rec = {
                    "brand_en":    brand_en,
                    "brand_cn":    brand_cn,
                    "name":        name,
                    "reg_num":     reg_raw,
                    "year":        year,
                    "row_idx":     row_idx,
                    "source_file": str(xlsx_path),
                    "editable":    (xlsx_path == Path(EXCEL_PATH)),
                    "ingredients": ingr,
                    "pdf_url":     pdf,
                    "poc_url":     poc,
                    "label_url":   label,
                    "is_special":  _is_special(reg_raw),
                    "is_imported": _is_imported(reg_raw),
                }

                missing = score_completeness(rec, img_map)
                if missing > 0:
                    rec["_missing"] = missing
                    records.append(rec)

        wb.close()

    # Sort: 2026 first → 2025 → 2024; within same year, more missing fields first
    records.sort(key=lambda r: (-r["year"], -r["_missing"]))
    return records


# ── PDF 渲染 → 图片 ───────────────────────────────────────────────────────
def render_pdf_image(rec: dict, img_map: dict) -> bool:
    """已有 pdf_url 的产品：尝试从本地下载缓存渲染首页为缩略图"""
    try:
        import fitz
        from PIL import Image as PILImage

        brand_en = rec["brand_en"]
        name     = rec["name"]
        prefix   = "特化--" if rec["is_special"] else ""
        # sanitize filename
        safe_name = re.sub(r'[\\/:*?"<>|\x00-\x1f]', "_", name)[:80]
        pdf_path  = DOWNLOAD_BASE / brand_en / f"{prefix}{safe_name}.pdf"

        if not pdf_path.exists() or pdf_path.stat().st_size == 0:
            return False

        out_dir = OUT_IMG_DIR / brand_en
        out_dir.mkdir(parents=True, exist_ok=True)
        out_jpg = out_dir / f"{safe_name}.jpg"

        doc = fitz.open(str(pdf_path))
        if doc.page_count == 0:
            return False
        page   = doc[0]
        rect   = page.rect
        scale  = IMG_MAX_PX / max(rect.width, rect.height, 1)
        scale  = max(min(scale, 3.0), 0.05)
        pix    = page.get_pixmap(matrix=fitz.Matrix(scale, scale))
        tmp    = io.BytesIO()
        tmp.write(pix.tobytes("png"))
        tmp.seek(0)
        img = PILImage.open(tmp).convert("RGB")
        w, h = img.size
        if max(w, h) > IMG_MAX_PX:
            r = IMG_MAX_PX / max(w, h)
            img = img.resize((max(1, int(w*r)), max(1, int(h*r))), PILImage.LANCZOS)
        img.save(str(out_jpg), "JPEG", quality=JPEG_Q, optimize=True)

        img_map.setdefault(brand_en, {})[name] = str(out_jpg).replace("\\", "/")
        log.info(f"    图片渲染成功: {out_jpg.name}")
        return True
    except Exception as e:
        log.debug(f"    PDF渲染失败 {rec['name']}: {e}")
        return False


# ── 从特殊注册 PDF 提取成分 ────────────────────────────────────────────────
def extract_ingredients_from_pdf(rec: dict) -> str:
    """特殊注册 PDF 是文字型，尝试提取全成分列表"""
    try:
        import fitz
        brand_en  = rec["brand_en"]
        name      = rec["name"]
        safe_name = re.sub(r'[\\/:*?"<>|\x00-\x1f]', "_", name)[:80]
        pdf_path  = DOWNLOAD_BASE / brand_en / f"特化--{safe_name}.pdf"
        if not pdf_path.exists() or pdf_path.stat().st_size == 0:
            return ""
        doc  = fitz.open(str(pdf_path))
        text = "\n".join(page.get_text() for page in doc)
        # 寻找全成分段落
        m = re.search(
            r"(?:全成分|成分表|配方成分|Ingredients?)[：:]\s*(\[?[^\n]{20,})",
            text, re.DOTALL
        )
        if m:
            raw = m.group(1).strip()
            # 取到下一个大段落前
            end = re.search(r"\n[^\n]{0,5}\n", raw)
            ingr = raw[:end.start()].strip() if end else raw[:600].strip()
            if ingr.startswith("["):
                ingr = ingr[1:]
            if ingr.endswith("]"):
                ingr = ingr[:-1]
            return ingr
    except Exception as e:
        log.debug(f"    成分提取失败 {rec['name']}: {e}")
    return ""


# ── 写回 Excel ────────────────────────────────────────────────────────────
def write_back(rec: dict, updates: dict, dry_run: bool) -> bool:
    """把 updates 里的列值写回当前 Excel（只操作 editable=True 的记录）。"""
    if not updates:
        return False
    if not rec["editable"]:
        log.info(f"    [跳过写回] {rec['name']} 在历史文件里，只写当前文件")
        return False
    if dry_run:
        log.info(f"    [dry-run] 会写入: {updates}")
        return True
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[rec["brand_en"]]
        for col, val in updates.items():
            ws.cell(row=rec["row_idx"], column=col, value=val)
        wb.save(EXCEL_PATH)
        log.info(f"    写入成功: {list(updates.keys())}")
        return True
    except Exception as e:
        log.error(f"    写入失败 {rec['name']}: {e}")
        return False


# ── 历史月份缺失检测与补全 ─────────────────────────────────────────────────
def _archive_name(year: int, month: int) -> str:
    """构造归档文件名，如 2025-06 → CI_List_Ada Jun'25.xlsx"""
    return f"CI_List_Ada {_MONTH_ABBR[month-1]}'{str(year)[2:]}.xlsx"


def detect_missing_months() -> list:
    """
    扫描全部 Excel 文件的实际数据，找出没有任何产品记录的月份。
    从 BACKFILL_START_YEAR 至上个月（不含当月，当月由月度任务负责）。
    2024年用 CI_List_Ada 2024.xlsx 统一覆盖，不单独按月检查。
    返回 [(year, month), ...] 按最近→最远排序。
    """
    from datetime import timedelta
    import re as _re

    today      = date.today()
    last_month = (today.replace(day=1) - timedelta(days=1))
    end_yr, end_mo = last_month.year, last_month.month

    # 统计各年月的产品数量
    month_counts: dict = {}   # "YYYY-MM" -> count
    date_pats = [
        _re.compile(r"(\d{4})-(\d{2})-\d{2}"),          # 2024-09-02 / ISO
        _re.compile(r"(\d{2})/\d{2}/(\d{4})"),           # 03/15/2026 / MM/DD/YYYY
    ]

    for xlsx_path in _all_excel_files():
        try:
            wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        except Exception:
            continue
        for sh in wb.sheetnames:
            if sh not in BRANDS:
                continue
            ws = wb[sh]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[COL_NAME - 1]:
                    continue
                # Check COL_DATE (col D) and COL_UPLOAD_DATE (col A)
                for col in [COL_DATE - 1, COL_UPLOAD_DATE - 1]:
                    val = row[col] if len(row) > col else None
                    s   = str(val or "").strip()
                    matched = False
                    for pat in date_pats:
                        m = pat.search(s)
                        if m:
                            if "/" in pat.pattern:      # MM/DD/YYYY
                                yr, mo = int(m.group(2)), int(m.group(1))
                            else:                       # YYYY-MM-DD
                                yr, mo = int(m.group(1)), int(m.group(2))
                            ym = f"{yr}-{mo:02d}"
                            month_counts[ym] = month_counts.get(ym, 0) + 1
                            matched = True
                            break
                    if matched:
                        break
        wb.close()

    # Determine which months are missing
    missing = []
    yr, mo = BACKFILL_START_YEAR, 1
    while (yr, mo) <= (end_yr, end_mo):
        if yr == 2024:
            # 2024 treated as one block — check if ANY 2024 months have data
            has_2024 = any(k.startswith("2024-") for k in month_counts)
            if not has_2024:
                missing.append((yr, mo))
            mo += 1
        else:
            ym = f"{yr}-{mo:02d}"
            if month_counts.get(ym, 0) == 0:
                missing.append((yr, mo))
            mo += 1
        if mo > 12:
            mo = 1
            yr += 1

    # Remove duplicates (2024) and sort most-recent first
    seen = set()
    deduped = []
    for item in missing:
        if item not in seen:
            seen.add(item)
            deduped.append(item)
    deduped.sort(reverse=True)
    return deduped


def _load_backfill_state() -> dict:
    if BACKFILL_STATE.exists():
        try:
            return json.loads(BACKFILL_STATE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"last_run": None, "attempted": [], "completed": []}


def _save_backfill_state(state: dict):
    BACKFILL_STATE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def _days_to_reach(year: int, month: int) -> int:
    """计算从今天到目标月份第一天所需的天数（+30 天覆盖全月）"""
    today   = date.today()
    target  = date(year, month, 1)
    delta   = (today - target).days + 30   # +30 覆盖整个月
    return max(delta, 40)


def run_backfill_if_needed(bebd_ok: bool, nmpa_ok: bool,
                           force: bool = False, dry_run: bool = False) -> bool:
    """
    检查是否需要运行历史月份补全（每7天一次）。
    force=True 则忽略7天限制。
    返回 True 表示执行了补全操作。
    """
    state = _load_backfill_state()

    # 检查是否到了运行时间（每7天）
    if not force:
        last = state.get("last_run")
        if last:
            days_since = (date.today() - date.fromisoformat(last)).days
            if days_since < 7:
                log.info(f"历史月份补全：距上次运行 {days_since} 天，未到7天间隔，跳过")
                return False

    missing = detect_missing_months()
    attempted  = set(state.get("attempted", []))
    completed  = set(state.get("completed", []))
    # 过滤掉已完成和连续失败超过3次的月份（避免无限重试）
    retry_counts = state.get("retry_counts", {})

    pending = [
        (yr, mo) for (yr, mo) in missing
        if f"{yr}-{mo:02d}" not in completed
        and retry_counts.get(f"{yr}-{mo:02d}", 0) < 3
    ]

    if not pending:
        log.info("历史月份补全：所有缺失月份已处理完毕或已达重试上限 🎉")
        return False

    # 取最近的缺失月份
    target_yr, target_mo = pending[0]
    target_key = f"{target_yr}-{target_mo:02d}"
    days_needed = _days_to_reach(target_yr, target_mo)

    log.info(f"\n{'='*60}")
    log.info(f"历史月份补全：目标 {target_key}（共 {len(pending)} 个缺失月份）")
    log.info(f"  使用 --days {days_needed} 回溯到 {target_yr}-{target_mo:02d}")
    log.info(f"  BEBD: {'✅' if bebd_ok else '❌'}  NMPA: {'✅' if nmpa_ok else '❌'}")

    if not bebd_ok:
        log.warning("  ⚠️  BEBD未登录，本次历史补全可能抓取数量偏少")

    if dry_run:
        log.info(f"  [dry-run] 会执行: python main.py --days {days_needed}")
        state["last_run"] = date.today().isoformat()
        _save_backfill_state(state)
        return True

    # 执行月度抓取
    cmd = [sys.executable, str(MAIN_SCRIPT), "--days", str(days_needed)]
    log.info(f"  执行: {' '.join(cmd)}")
    start_t = datetime.now()
    result  = subprocess.run(
        cmd,
        cwd=str(WORKSPACE),
        capture_output=False,
        timeout=7200   # 最多2小时
    )
    elapsed = (datetime.now() - start_t).seconds // 60

    # 更新状态
    state["last_run"]   = date.today().isoformat()
    state["attempted"]  = list(set(state.get("attempted", [])) | {target_key})
    retry_counts[target_key] = retry_counts.get(target_key, 0) + 1
    state["retry_counts"] = retry_counts

    if result.returncode == 0:
        # 检查是否产生了归档文件（由 main.py 的 step 1 写入月度Excel）
        base      = Path(EXCEL_PATH).parent
        arch_name = _archive_name(target_yr, target_mo)
        if (base / arch_name).exists() or target_yr == 2024:
            state["completed"] = list(set(state.get("completed", [])) | {target_key})
            log.info(f"  ✅ {target_key} 补全完成（耗时 {elapsed} 分钟）")
        else:
            log.warning(f"  ⚠️ 抓取完成但未找到归档文件 {arch_name}，下次重试")
    else:
        log.error(f"  ❌ 抓取失败（exit={result.returncode}），下次重试")

    _save_backfill_state(state)

    # 剩余缺失月份摘要
    remaining = len(pending) - (1 if result.returncode == 0 else 0)
    log.info(f"  剩余缺失月份: {remaining} 个，预计 {remaining * 7} 天内补全")
    log.info("=" * 60)
    return True


# ── 主流程 ────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit",    type=int,  default=10, help="每次最多处理 N 个产品（默认10）")
    ap.add_argument("--dry-run",  action="store_true", help="只统计，不写入不渲染")
    ap.add_argument("--backfill", action="store_true", help="立即运行历史月份补全（忽略7天间隔）")
    ap.add_argument("--check",   action="store_true", help="只做连通性检查后退出")
    args = ap.parse_args()

    log.info("=" * 60)
    log.info(f"每日增量补全任务  {datetime.today().strftime('%Y-%m-%d %H:%M')}")
    log.info("=" * 60)

    img_map: dict = json.loads(MAP_JSON.read_text(encoding="utf-8")) if MAP_JSON.exists() else {}

    # ── 连通性检查 ──────────────────────────────────────────────────────
    bebd_ok = nmpa_ok = False
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage",
                      "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"]
            )
            ctx  = browser.new_context(ignore_https_errors=True)
            page = ctx.new_page()
            log.info("连通性检查...")
            bebd_ok = check_bebd(page)
            nmpa_ok = check_nmpa(page)
            browser.close()
    except Exception as e:
        log.warning(f"连通性检查异常: {e}")

    if args.check:
        log.info(f"BEBD登录: {'✅' if bebd_ok else '❌'}  NMPA hzpba: {'✅' if nmpa_ok else '❌'}")
        return

    # ── 统计缺失 ────────────────────────────────────────────────────────
    log.info("扫描所有 Excel 文件寻找不完整产品...")
    records = load_incomplete_records(img_map)
    log.info(f"共 {len(records)} 个不完整产品（按 2026→2025→2024 排序）")

    by_year: dict = {}
    for r in records:
        by_year.setdefault(r["year"], 0)
        by_year[r["year"]] += 1
    for yr in sorted(by_year, reverse=True):
        log.info(f"  {yr}: {by_year[yr]} 个")

    todo = records[:args.limit]
    log.info(f"\n本次处理前 {len(todo)} 个:")

    img_map_dirty = False
    processed = ok = 0

    for rec in todo:
        name = rec["name"]
        log.info(f"\n[{rec['brand_en']}] {name}  (year={rec['year']}, missing={rec['_missing']})")
        updates: dict = {}

        # 1. 本地PDF → 图片渲染（不需要网络）
        if not _has_image(rec["brand_en"], name, img_map):
            if render_pdf_image(rec, img_map):
                img_map_dirty = True

        # 2. 特殊注册 → 从 PDF 提取成分
        if not rec["ingredients"] and rec["is_special"] and not args.dry_run:
            ingr = extract_ingredients_from_pdf(rec)
            if ingr:
                updates[COL_INGREDIENTS] = ingr
                log.info(f"    提取成分: {ingr[:60]}...")

        # 3. BEBD / NMPA 在线补全（需要连通性）
        if (not rec["pdf_url"] or not rec["poc_url"]) and (bebd_ok or nmpa_ok):
            try:
                from playwright.sync_api import sync_playwright
                from module2_nmpa import search_hzpba
                from module1_bebd import get_nmpa_special_pdf_url

                with sync_playwright() as pw:
                    browser = pw.chromium.launch(
                        headless=True,
                        args=["--no-sandbox", "--disable-dev-shm-usage",
                              "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"]
                    )
                    ctx  = browser.new_context(ignore_https_errors=True)
                    page = ctx.new_page()

                    if not rec["pdf_url"] and not rec["is_special"] and nmpa_ok:
                        # 普通备案 → hzpba
                        log.info(f"    尝试 hzpba 查询 pdf_url ...")
                        pdf_url = search_hzpba(
                            page, rec["reg_num"], name,
                            is_imported=rec["is_imported"]
                        )
                        if pdf_url:
                            updates[COL_PDF_URL] = pdf_url
                            rec["pdf_url"] = pdf_url
                            log.info(f"    pdf_url: {pdf_url[:70]}")

                    if not rec["pdf_url"] and rec["is_special"] and nmpa_ok:
                        # 特殊注册 → nmpa.gov.cn/datasearch
                        log.info(f"    尝试 nmpa datasearch 查询 ...")
                        try:
                            special_url = get_nmpa_special_pdf_url(
                                page, rec["reg_num"],
                                is_imported=rec["is_imported"]
                            )
                            if special_url:
                                updates[COL_PDF_URL] = special_url
                                rec["pdf_url"] = special_url
                                log.info(f"    特殊注册 pdf_url: {special_url[:70]}")
                        except Exception as e2:
                            log.debug(f"    nmpa datasearch 失败: {e2}")

                    browser.close()
            except Exception as e:
                log.warning(f"    在线补全失败: {e}")

        # 4. 如果刚拿到 pdf_url，尝试下载 + 渲染图片
        if updates.get(COL_PDF_URL) and not _has_image(rec["brand_en"], name, img_map):
            try:
                from module3_download import download_file, sanitize
                brand_dir = DOWNLOAD_BASE / rec["brand_en"]
                brand_dir.mkdir(parents=True, exist_ok=True)
                prefix    = "特化--" if rec["is_special"] else ""
                safe_name = sanitize(name)
                local_pdf = brand_dir / f"{prefix}{safe_name}.pdf"
                if not args.dry_run:
                    download_file(updates[COL_PDF_URL], local_pdf)
                    if render_pdf_image(rec, img_map):
                        img_map_dirty = True
            except Exception as e:
                log.debug(f"    PDF下载/渲染失败: {e}")

        # 5. 写回 Excel
        if write_back(rec, updates, args.dry_run):
            ok += 1
        processed += 1

    # ── 保存 image_map ───────────────────────────────────────────────────
    if img_map_dirty and not args.dry_run:
        MAP_JSON.write_text(json.dumps(img_map, ensure_ascii=False, indent=2), encoding="utf-8")
        log.info("\nimage_map.json 已更新")

    # ── 历史缺失月份摘要 + 补全（每7天一次，或 --backfill 强制）──────────
    missing_months = detect_missing_months()
    state = _load_backfill_state()
    completed = set(state.get("completed", []))
    pending_months = [
        (yr, mo) for yr, mo in missing_months
        if f"{yr}-{mo:02d}" not in completed
    ]
    if pending_months:
        log.info(f"\n缺失月份 ({len(pending_months)} 个，最近→最远):")
        for yr, mo in pending_months[:10]:
            days = _days_to_reach(yr, mo)
            log.info(f"  {yr}-{mo:02d}  (需 --days {days} 约 {days//30} 个月前)")
        if len(pending_months) > 10:
            log.info(f"  ... 还有 {len(pending_months)-10} 个更早的月份")
        log.info(f"  每周补全1个月，预计 {len(pending_months)} 周内完成")

        run_backfill_if_needed(
            bebd_ok=bebd_ok,
            nmpa_ok=nmpa_ok,
            force=getattr(args, "backfill", False),
            dry_run=args.dry_run,
        )
    else:
        log.info("\n✅ 所有历史月份数据均已补全")

    log.info(f"\n{'='*60}")
    log.info(f"完成: 处理 {processed} 个, 写入更新 {ok} 个")
    log.info(f"BEBD: {'✅' if bebd_ok else '❌'}  NMPA: {'✅' if nmpa_ok else '❌'}")
    log.info(f"剩余不完整产品: {len(records) - processed} 个（约 {max(0,(len(records)-processed+9)//10)} 天可补全）")
    log.info(f"缺失历史月份: {len(pending_months)} 个（约 {len(pending_months)} 周可补全，每周一个）")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
