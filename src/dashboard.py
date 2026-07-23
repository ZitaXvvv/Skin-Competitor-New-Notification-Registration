"""
CI New SKU Dashboard — Streamlit 前端
月历视图：每格显示产品名（中文 + 英文）、功效、Artwork PDF、mini-POC 链接

运行：
    streamlit run src/dashboard.py
"""

import base64
import json
import os
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import streamlit as st

sys.path.insert(0, str(Path(__file__).parent))
from config import (
    BRANDS,
    COL_CATEGORY,
    COL_DATE,
    COL_EFFECT,
    COL_INGREDIENTS,
    COL_LABEL_URL,
    COL_NAME,
    COL_PDF_URL,
    COL_POC_URL,
    COL_REG_NUM,
    COL_UPLOAD_DATE,
    DOWNLOAD_BASE,
    EXCEL_PATH,
    LOG_DIR,
)
from auth import verify_login

# ─────────────────────────────────────────────
# 页面配置
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="CI New SKU · Competitor Intelligence",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="collapsed",
)

TRANSLATION_CACHE = Path(__file__).parent / "translations_cache.json"
IMAGE_MAP_JSON    = Path(__file__).parent.parent / "res" / "product_images" / "image_map.json"


@st.cache_resource
def load_image_map() -> dict[str, dict[str, str]]:
    """加载产品名→图片路径映射，并预先 base64 编码图片"""
    if not IMAGE_MAP_JSON.exists():
        return {}
    raw: dict[str, dict[str, str]] = json.loads(IMAGE_MAP_JSON.read_text(encoding="utf-8"))
    result: dict[str, dict[str, str]] = {}   # brand_en → {prod_name: data_uri}
    for brand, mapping in raw.items():
        result[brand] = {}
        for name, path in mapping.items():
            p = Path(path)
            if p.exists():
                try:
                    mime = "image/jpeg" if p.suffix.lower() in (".jpg", ".jpeg") else "image/png"
                    b64 = base64.b64encode(p.read_bytes()).decode()
                    uri = f"data:{mime};base64,{b64}"
                    # 拆分拼接名（如"名称A名称B"→分别存两份 uri）
                    for part in _split_img_key(name):
                        result[brand][part] = uri
                except Exception:
                    pass
    return result


# 品牌中文名列表（用于从 image key 中识别/剥离品牌前缀）
_ZH_BRANDS = [
    "珀莱雅","谷雨","欧诗漫","兰蔻","欧莱雅","雅诗兰黛",
    "修丽可","百雀羚","韩束","自然堂","薇诺娜","妮维雅","资生堂","科颜氏","契尔氏",
]

def _split_img_key(name: str) -> list[str]:
    """
    把拼接的图片 key 拆开，并为每个子名生成带/不带品牌的变体。
    例：'珀莱雅双抗焕亮清透水珀莱雅双抗焕白净亮清透水'
      → ['珀莱雅双抗焕亮清透水', '双抗焕亮清透水',
          '珀莱雅双抗焕白净亮清透水', '双抗焕白净亮清透水']
    """
    import re
    parts = [name]
    # 在已知品牌名前拆分（同一品牌出现两次 = 两个名字拼一起）
    for zh in _ZH_BRANDS:
        new_parts = []
        for p in parts:
            # 找第二次出现的品牌名，在那里切割
            idx = p.find(zh, len(zh))  # 从品牌名长度后开始找
            if idx > 0:
                new_parts.append(p[:idx])
                new_parts.append(p[idx:])
            else:
                new_parts.append(p)
        parts = new_parts

    result = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        result.append(p)
        # 去品牌前缀的变体
        for zh in _ZH_BRANDS:
            if p.startswith(zh):
                short = p[len(zh):]
                if len(short) >= 4:
                    result.append(short)
                break
    return list(dict.fromkeys(result))  # 保序去重


def find_prod_img(brand_en: str, prod_name: str, img_map: dict) -> str:
    """
    在 img_map[brand_en] 中查找 prod_name 对应的图片 URI。
    顺序：精确 → 去品牌前缀 → 图片key是产品名子串 → 产品名是图片key子串
    """
    brand_imgs = img_map.get(brand_en, {})
    if not brand_imgs:
        return ""

    # 1. 精确匹配
    if prod_name in brand_imgs:
        return brand_imgs[prod_name]

    # 2. 去品牌前缀后精确匹配
    short_name = prod_name
    for zh in _ZH_BRANDS:
        if prod_name.startswith(zh):
            short_name = prod_name[len(zh):]
            break
    if short_name != prod_name and short_name in brand_imgs:
        return brand_imgs[short_name]

    # 3. 图片 key（或其去品牌版）是产品名的子串（最长优先）
    candidates = []
    for img_key, uri in brand_imgs.items():
        # img_key 去品牌前缀
        img_short = img_key
        for zh in _ZH_BRANDS:
            if img_key.startswith(zh):
                img_short = img_key[len(zh):]
                break
        # 用去品牌的短名做子串匹配
        if len(img_short) >= 4 and img_short in prod_name:
            candidates.append((len(img_short), uri))
        elif len(img_short) >= 4 and img_short in short_name:
            candidates.append((len(img_short), uri))
        # 反向：产品名去品牌后是图片key的子串
        elif len(short_name) >= 4 and short_name in img_short:
            candidates.append((len(short_name), uri))

    if candidates:
        candidates.sort(key=lambda x: -x[0])  # 最长匹配优先
        return candidates[0][1]

    return ""

# ─────────────────────────────────────────────
# 本地翻译：美妆专用词典（零网络依赖）
# ─────────────────────────────────────────────

# 品牌中英文映射
_BRAND_MAP = {
    "珀莱雅": "PROYA", "谷雨": "Guyu", "欧诗漫": "OSM",
    "科颜氏": "Kiehl's", "欧莱雅": "L'Oréal", "雅诗兰黛": "Estée Lauder",
    "修丽可": "SkinCeuticals", "兰蔻": "Lancôme", "百雀羚": "Pechoin",
    "自然堂": "Chando", "薇诺娜": "Winona", "韩束": "Kans",
    "娇韵诗": "Clarins", "契尔氏": "Kiehl's",
}

# 产品剂型
_FORM_MAP = {
    "精华液": "Serum", "精华": "Serum", "面霜": "Face Cream", "眼霜": "Eye Cream",
    "乳液": "Lotion", "水乳": "Toner & Lotion", "乳": "Milk/Lotion",
    "面膜": "Sheet Mask", "睡眠面膜": "Sleeping Mask",
    "慕斯": "Mousse", "泡沫": "Foam",
    "洁面": "Cleanser", "卸妆": "Makeup Remover",
    "防晒乳": "Sunscreen Lotion", "防晒霜": "Sunscreen Cream",
    "防晒": "Sunscreen", "隔离": "Primer",
    "喷雾": "Mist", "喷": "Mist", "露": "Dew",
    "油": "Oil", "霜": "Cream", "啫喱": "Gel",
    "次抛": "Ampoule", "安瓶": "Ampoule",
    "精粹": "Essence", "素": "Essence",
    "唇膏": "Lip Balm", "唇霜": "Lip Cream",
    "爽肤水": "Toner", "化妆水": "Toner",
    "冻膜": "Jelly Mask", "凝露": "Gel Serum",
    "凝霜": "Gel Cream", "凝冻": "Gel",
}

# 功效关键词
_BENEFIT_MAP = {
    "水润": "Hydrating", "保湿": "Moisturizing", "水光": "Luminous",
    "滋润": "Nourishing", "锁水": "Moisture-Lock",
    "美白": "Brightening", "焕白": "Radiance", "亮白": "Whitening",
    "抗皱": "Anti-Wrinkle", "淡纹": "Wrinkle Reduction",
    "紧致": "Firming", "塑颜": "Sculpting",
    "修护": "Repairing", "修复": "Recovery",
    "舒缓": "Soothing", "镇静": "Calming",
    "控油": "Oil Control", "祛痘": "Acne-Clearing",
    "去角质": "Exfoliating", "焕肤": "Renewal",
    "提亮": "Illuminating", "焕亮": "Brightening",
    "双抗": "Dual Antioxidant", "抗氧": "Antioxidant",
    "源力": "Energy-Boost", "赋能": "Energizing",
    "胶原": "Collagen", "玻尿酸": "Hyaluronic Acid",
    "烟酰胺": "Niacinamide", "视黄醇": "Retinol",
    "特护": "Intensive Care", "舒敏": "Sensitive-Soothing",
    "净澈": "Purifying", "净透": "Clarifying",
    "凝时": "Time Freeze", "冻龄": "Age-Freeze",
    "极润": "Ultra-Moist", "嘭盈": "Plumping",
    "缎光": "Satin Glow", "盈润": "Dewy",
}


def translate_local(zh_name: str) -> str:
    """本地美妆词典翻译，不需要网络。"""
    if not zh_name:
        return ""

    result = zh_name

    # 1. 替换品牌名
    for zh, en in _BRAND_MAP.items():
        result = result.replace(zh, en)

    # 2. 替换剂型（从长到短，避免截断）
    for zh, en in sorted(_FORM_MAP.items(), key=lambda x: -len(x[0])):
        result = result.replace(zh, f" {en} ")

    # 3. 替换功效关键词
    for zh, en in sorted(_BENEFIT_MAP.items(), key=lambda x: -len(x[0])):
        result = result.replace(zh, f" {en} ")

    # 4. 清理多余空格，保留首尾
    parts = [p.strip() for p in result.split() if p.strip()]
    translated = " ".join(parts)

    # 5. 若仍有汉字（未覆盖词），保留原文追加
    import unicodedata
    has_cjk = any(unicodedata.category(c) in ("Lo",) for c in translated)
    return translated if not has_cjk or translated != zh_name else zh_name


@st.cache_resource
def get_translation_cache() -> dict:
    if TRANSLATION_CACHE.exists():
        with open(TRANSLATION_CACHE, encoding="utf-8") as f:
            return json.load(f)
    return {}


def translate_batch(names: list[str], cache: dict, timeout_sec: int = 8) -> dict:
    """批量翻译：优先用本地词典，无需网络。"""
    for name in names:
        if name and name not in cache:
            cache[name] = translate_local(name)
    # 持久化（本地翻译很快，直接保存）
    try:
        with open(TRANSLATION_CACHE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    return cache


# ─────────────────────────────────────────────
# 数据加载（读取当前 + 所有历史月份 Excel，跨文件去重）
# ─────────────────────────────────────────────
import re as _re_global
_REG_PAT_GLOBAL = _re_global.compile(
    r"(妆网备字|国妆备字|国妆特字|国妆特进字|国妆备进字|卫妆特字)"
)

# 前端不展示：唇部类产品 & 男士类产品
# 名称关键词：含"唇"字的都排除（涵盖唇油/唇蜜/唇露/唇膏/唇彩/唇釉/润唇等）
_EXCLUDE_NAME_PAT = _re_global.compile(
    r"唇|口红|男士|男仕"
)
# 类目关键词：category 列含"唇"字也排除
_EXCLUDE_CAT_PAT = _re_global.compile(r"唇")


def _excel_files_to_load() -> list[Path]:
    """返回所有 CI_List_Ada*.xlsx 路径，当前文件排最后（优先级最高）"""
    base_dir = Path(EXCEL_PATH).parent
    historical = sorted(base_dir.glob("CI_List_Ada *.xlsx"))
    current = Path(EXCEL_PATH)
    # 排除已在 historical 里的当前文件（名字不同）
    result = [f for f in historical if f != current]
    if current.exists():
        result.append(current)
    return result


def _parse_notif_date(notif_raw, upload_raw):
    from datetime import date as _date, datetime as _datetime
    for val in [notif_raw, upload_raw]:
        if val is None:
            continue
        if hasattr(val, "date") and callable(val.date):
            return val.date()
        if isinstance(val, _date):
            return val
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return _datetime.strptime(str(val).strip(), fmt).date()
            except ValueError:
                pass
    return None


def _load_one_excel(path: Path, seen_regs: set, records: list):
    """从单个 Excel 读取所有品牌数据，去重后追加到 records"""
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception:
        return
    for brand_en, brand_cn in BRANDS.items():
        if brand_en not in wb.sheetnames:
            continue
        ws = wb[brand_en]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 5:
                continue

            def g(col):
                return row[col - 1] if len(row) >= col else None

            name_raw   = g(COL_NAME)
            effect_raw = g(COL_EFFECT)
            notif_raw  = g(COL_DATE)
            upload_raw = g(COL_UPLOAD_DATE)
            ingr_raw   = g(COL_INGREDIENTS)
            pdf_raw    = g(COL_PDF_URL)
            label_raw  = g(COL_LABEL_URL)
            poc_raw    = g(COL_POC_URL)

            if not name_raw:
                continue

            # 排除唇部/男士类产品，不在前端展示
            # 同时检查 category 列（部分产品名称无"唇"字但类目是唇部）
            category_raw = g(COL_CATEGORY)
            if _EXCLUDE_NAME_PAT.search(str(name_raw)):
                continue
            if category_raw and _EXCLUDE_CAT_PAT.search(str(category_raw)):
                continue

            notif_date = _parse_notif_date(notif_raw, upload_raw)
            if not notif_date:
                continue

            # 找备案号（任意列）
            reg_str = ""
            for cell in row:
                s = str(cell or "").strip()
                if s and _REG_PAT_GLOBAL.search(s):
                    reg_str = s
                    break

            # 跨文件去重（优先保留当前文件的数据，historical 已先加入）
            if reg_str and reg_str in seen_regs:
                continue
            if reg_str:
                seen_regs.add(reg_str)

            reg_type = ""
            if _re_global.search(r"国妆特字|国妆特进字|卫妆特字", reg_str):
                reg_type = "特殊注册"
            elif _re_global.search(r"妆网备字|国妆备字|国妆备进字", reg_str):
                reg_type = "普通备案"

            def clean(v):
                s = str(v or "").strip()
                return s if s not in ("None", "NA", "") else ""

            records.append({
                "brand_en":    brand_en,
                "brand_cn":    brand_cn,
                "name":        str(name_raw).strip(),
                "effect":      clean(effect_raw),
                "ingredients": clean(ingr_raw),
                "notif_date":  notif_date,
                "year":        notif_date.year,
                "month":       notif_date.month,
                "year_month":  notif_date.strftime("%Y-%m"),
                "reg_num":     reg_str,
                "reg_type":    reg_type,
                "pdf_url":     clean(pdf_raw),
                "label_url":   clean(label_raw),
                "poc_url":     clean(poc_raw),
                "source_file": path.name,
                "editable":    (path == Path(EXCEL_PATH)),
                "row_idx":     row_idx,
            })
    wb.close()


@st.cache_data(ttl=1800, show_spinner="⏳ 读取数据…")
def load_data() -> list[dict]:
    files = _excel_files_to_load()
    if not files:
        return []
    records: list[dict] = []
    seen_regs: set[str] = set()
    for f in files:
        _load_one_excel(f, seen_regs, records)
    return records
    return records


# ─────────────────────────────────────────────
# CSS（现代设计风格）
# ─────────────────────────────────────────────
GLOBAL_CSS = """
<style>
  /* ── 全局字体 ── */
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
  html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

  /* ── 隐藏 Streamlit 默认元素 ── */
  #MainMenu, footer, header { visibility: hidden; }
  .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }

  /* ── multiselect 选中标签：去掉红色，改为中性蓝灰色 ── */
  [data-baseweb="tag"] {
    background-color: #e8edf5 !important;
    border: 1px solid #c5cfe0 !important;
    border-radius: 6px !important;
  }
  [data-baseweb="tag"] span { color: #1a2b4a !important; font-weight: 500; font-size: 12px; }
  [data-baseweb="tag"] button {
    color: #5f7089 !important;
    background: transparent !important;
  }
  [data-baseweb="tag"] button:hover { color: #1a2b4a !important; }
  /* multiselect 输入框 展开时高亮 */
  [data-baseweb="select"] [data-testid="stMultiSelectInput"]:focus-within {
    border-color: #3d7bd9 !important;
    box-shadow: 0 0 0 2px rgba(61,123,217,.15) !important;
  }

  /* ── 顶部 Hero ── */
  .hero {
    background: linear-gradient(135deg, #0f2027, #203a43, #2c5364);
    border-radius: 16px;
    padding: 28px 36px;
    margin-bottom: 24px;
    color: white;
  }
  .hero h1 { font-size: 28px; font-weight: 700; margin: 0 0 4px 0; letter-spacing: -0.5px; }
  .hero p  { font-size: 14px; color: rgba(255,255,255,0.65); margin: 0; }

  /* ── 统计卡片 ── */
  .stat-row { display: flex; gap: 16px; margin-bottom: 24px; }
  .stat-card {
    flex: 1;
    background: white;
    border: 1px solid #e8ecf0;
    border-radius: 12px;
    padding: 18px 22px;
    box-shadow: 0 1px 4px rgba(0,0,0,.06);
  }
  .stat-card .label { font-size: 11px; font-weight: 600; color: #8898aa;
                      text-transform: uppercase; letter-spacing: .6px; }
  .stat-card .value { font-size: 32px; font-weight: 700; color: #1a2b4a;
                      line-height: 1.1; margin-top: 4px; }
  .stat-card.blue  .value { color: #1565c0; }
  .stat-card.green .value { color: #2e7d32; }
  .stat-card.red   .value { color: #c62828; }

  /* ── 日历表格 ── */
  .cal-wrap { border-radius: 12px;
              box-shadow: 0 2px 12px rgba(0,0,0,.08); }
  .cal-table { border-collapse: collapse; width: 100%; table-layout: fixed;
               font-size: 12px; background: white; }

  /* 表头 */
  .cal-th {
    background: #1a2b4a;
    color: white;
    text-align: center;
    padding: 10px 6px;
    font-weight: 600;
    font-size: 11px;
    letter-spacing: .4px;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
  }
  .cal-th-brand {
    background: #1a2b4a;
    color: white;
    padding: 10px 14px;
    font-weight: 600;
    font-size: 11px;
    text-align: left;
    width: 88px;
    position: sticky;
    left: 0;
    z-index: 2;
  }

  /* 品牌列 */
  .cal-brand {
    background: #f8f9fb;
    border-right: 2px solid #e2e8f0;
    padding: 12px 14px;
    font-weight: 700;
    font-size: 12px;
    color: #1a2b4a;
    vertical-align: top;
    width: 88px;
    position: sticky;
    left: 0;
    z-index: 1;
  }
  .cal-brand small { display: block; color: #8898aa;
                     font-weight: 400; font-size: 10px; margin-top: 2px; }

  /* 数据格 */
  .cal-cell {
    border: 1px solid #f0f2f5;
    padding: 6px;
    vertical-align: top;
    background: white;
    overflow: hidden;
  }
  .cal-cell:hover { background: #fafbff; }

  /* 每格：卡片自适应换行——不再固定宽度，根据实际格子宽度自动排列，避免整个表格溢出需要横向拖动 */
  .cell-grid {
    display: flex;
    flex-direction: row;
    flex-wrap: wrap;
    width: 100%;
    gap: 6px;
  }

  /* 产品卡片：始终每行3张，宽度响应格子实际宽度 (gap=6px, 3列: (100%-2*6)/3) */
  .prod-card {
    background: white;
    border: 1px solid #e8ecf1;
    border-top: 3px solid #1565c0;
    border-radius: 7px;
    box-sizing: border-box;
    overflow: hidden;
    box-shadow: 0 1px 3px rgba(0,0,0,.04);
    flex: 0 0 calc((100% - 12px) / 3);
    width: calc((100% - 12px) / 3);
    min-height: 90px;
    display: flex;
    flex-direction: column;
  }
  .prod-card:hover { box-shadow: 0 3px 10px rgba(0,0,0,.10); }
  .prod-card.special { border-top-color: #c62828; }

  /* 产品图片 */
  .prod-img-wrap {
    width: 100%;
    aspect-ratio: 1;
    overflow: hidden;
    background: #f8f9fb;
    display: flex;
    align-items: center;
    justify-content: center;
    flex-shrink: 0;
  }
  .prod-img-wrap img {
    width: 100%;
    height: 100%;
    object-fit: contain;
  }
  .prod-img-placeholder {
    width: 100%;
    aspect-ratio: 1;
    background: #f4f6f9;
    display: flex;
    align-items: center;
    justify-content: center;
    color: #cdd5df;
    flex-shrink: 0;
  }
  .prod-img-placeholder svg { width: 30%; height: 30%; }

  /* <details> 折叠 */
  .prod-card details { margin: 0; }
  .prod-card summary {
    cursor: pointer; list-style: none;
    padding: 7px 8px;
    display: flex; flex-direction: column; gap: 3px;
  }
  .prod-card summary::-webkit-details-marker { display: none; }
  .prod-card summary:hover { background: #f5f8ff; }
  .sum-row { display: flex; align-items: flex-start; justify-content: space-between; gap: 4px; }
  .expand-hint { color: #aab4be; font-size: 9px; flex-shrink: 0; }
  details[open] .expand-hint::after { content: '▲'; }
  details:not([open]) .expand-hint::after { content: '▼'; }
  .prod-body { padding: 0 8px 8px; border-top: 1px solid #f0f2f5; }

  .prod-name { font-weight: 600; color: #1a2b4a; font-size: 11px;
               line-height: 1.3; flex: 1; }
  .prod-en   { color: #5f7089; font-size: 10px; margin: 5px 0 3px;
               font-style: italic; }
  .prod-eff  { color: #637382; font-size: 10px; margin-bottom: 5px;
               display: -webkit-box; -webkit-line-clamp: 2;
               -webkit-box-orient: vertical; overflow: hidden; }
  .prod-meta { display: flex; align-items: center; gap: 5px;
               margin-bottom: 5px; flex-wrap: wrap; }

  /* 徽章 */
  .badge {
    font-size: 9px; font-weight: 700; padding: 2px 7px;
    border-radius: 20px; white-space: nowrap;
  }
  .badge-n  { background: #e8f5e9; color: #2e7d32; border: 1px solid #c8e6c9; }
  .badge-s  { background: #ffebee; color: #c62828; border: 1px solid #ffcdd2; }

  /* 按钮行 */
  .btn-row  { display: flex; gap: 5px; flex-wrap: wrap; margin-top: 4px; }
  .btn {
    font-size: 10px; font-weight: 600;
    padding: 3px 9px; border-radius: 5px;
    text-decoration: none; display: inline-flex;
    align-items: center; gap: 3px; border: none;
    cursor: pointer; white-space: nowrap;
  }
  .btn-art { background: #1565c0; color: white; }
  .btn-art:hover { background: #0d47a1; color: white; }
  .btn-lbl { background: #6a1b9a; color: white; }
  .btn-poc { background: #2e7d32; color: white; }

  /* 月份徽章 */
  .month-pill { background: #3d7bd9; color: white; border-radius: 12px;
                padding: 1px 7px; font-size: 10px; margin-left: 4px;
                font-weight: 600; }

  /* 筛选栏 */
  .filter-bar {
    background: white;
    border: 1px solid #e8ecf0;
    border-radius: 12px;
    padding: 16px 20px;
    margin-bottom: 20px;
    box-shadow: 0 1px 4px rgba(0,0,0,.05);
  }
</style>
"""


def product_card_html(prod: dict, en_name: str, img_uri: str = "") -> str:
    is_special = prod["reg_type"] == "特殊注册"
    badge = (f'<span class="badge badge-s">特殊</span>'
             if is_special else f'<span class="badge badge-n">备案</span>')

    # 图片区域
    if img_uri:
        img_block = f'<div class="prod-img-wrap"><img src="{img_uri}" alt=""></div>'
    else:
        img_block = ('<div class="prod-img-placeholder">'
                     '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5">'
                     '<path d="M9 2h6M10 2v3.2c0 .3-.1.6-.3.8L7.6 8.7c-.4.4-.6 1-.6 1.6V20a1 1 0 0 0 1 1h8a1 1 0 0 0 1-1v-9.7c0-.6-.2-1.2-.6-1.6l-2.1-2.7a1.3 1.3 0 0 1-.3-.8V2" '
                     'stroke-linecap="round" stroke-linejoin="round"/></svg></div>')

    card_cls = "prod-card special" if is_special else "prod-card"
    name_disp = prod["name"][:28]
    pid = prod.get("_pid", "")
    drag_attrs = (f'draggable="true" data-pid="{pid}" ondragstart="cmpDragStart(event)"'
                  if pid else "")

    # Artwork/POC 按钮
    btns = ""
    if prod["pdf_url"]:
        btns += f'<a href="{prod["pdf_url"]}" target="_blank" class="btn btn-art" title="NMPA PDF">&#x1F5BC;</a>'
    if prod["label_url"]:
        btns += f'<a href="{prod["label_url"]}" target="_blank" class="btn btn-lbl" title="Label">&#x1F3F7;</a>'
    if prod["poc_url"]:
        btns += f'<a href="{prod["poc_url"]}" target="_blank" class="btn btn-poc" title="NMPA \u4ea7\u54c1\u8be6\u60c5">&#x1F9EA;</a>'
    btn_block = f'<div class="btn-row" style="padding:3px 5px 4px;flex-wrap:wrap">{btns}</div>' if btns else ""

    return f"""<div class="{card_cls}" {drag_attrs}>
  {img_block}
  <div style="padding:5px 6px 0">
    <div class="prod-name" style="font-size:10px;line-height:1.3">{name_disp}</div>
    <div style="margin-top:3px">{badge}</div>
  </div>
  {btn_block}
</div>"""


def _quarter_key(year: int, month: int) -> str:
    q = (month - 1) // 3 + 1
    return f"{year}-Q{q}"


def build_calendar_html(records, selected_brands, months, trans_cache, img_map) -> str:
    grouped: dict[str, dict[str, list]] = {}
    for r in records:
        if r["brand_en"] not in selected_brands:
            continue
        qk = _quarter_key(r["year"], r["month"])
        grouped.setdefault(r["brand_en"], {}).setdefault(qk, []).append(r)

    if not grouped:
        return "<p style='color:#888;padding:20px'>当前筛选条件下无数据</p>"

    html = '<div class="cal-wrap"><table class="cal-table"><thead><tr>'
    html += '<th class="cal-th-brand">品牌</th>'
    for ym, label in months:
        total = sum(len(grouped.get(b, {}).get(ym, [])) for b in selected_brands)
        pill = f'<span class="month-pill">{total}</span>' if total else ""
        html += f'<th class="cal-th">{label}{pill}</th>'
    html += "</tr></thead><tbody>"

    for brand_en in selected_brands:
        brand_data = grouped.get(brand_en)
        if not brand_data:
            continue
        brand_cn = BRANDS[brand_en]
        brand_imgs = img_map.get(brand_en, {})
        html += f'<tr><td class="cal-brand">{brand_en}<small>{brand_cn}</small></td>'
        for ym, _ in months:
            prods = brand_data.get(ym, [])
            html += '<td class="cal-cell"><div class="cell-grid">'
            for p in prods:
                en = trans_cache.get(p["name"], "")
                img_uri = find_prod_img(brand_en, p["name"], img_map)
                html += product_card_html(p, en, img_uri)
            html += "</div></td>"
        html += "</tr>"

    html += "</tbody></table></div>"
    return html


# ─────────────────────────────────────────────
# 成分中英文词典（INCI / 通用名）
# ─────────────────────────────────────────────
_INGR_ZH2EN: dict[str, str] = {
    # 基础
    "水": "Water (Aqua)", "去离子水": "Purified Water", "蒸馏水": "Distilled Water",
    "甘油": "Glycerin", "丙二醇": "Propylene Glycol", "丁二醇": "Butylene Glycol",
    "戊二醇": "Pentylene Glycol", "己二醇": "Hexylene Glycol",
    "1,3-丙二醇": "1,3-Propanediol", "双丙甘醇": "Dipropylene Glycol",
    # 活性成分
    "透明质酸钠": "Sodium Hyaluronate", "玻尿酸": "Hyaluronic Acid",
    "烟酰胺": "Niacinamide", "视黄醇": "Retinol", "视黄醇棕榈酸酯": "Retinyl Palmitate",
    "抗坏血酸": "Ascorbic Acid (Vit.C)", "抗坏血酸葡糖苷": "Ascorbyl Glucoside",
    "3-邻-乙基抗坏血酸": "Ethyl Ascorbic Acid", "维生素C": "Vitamin C",
    "生育酚": "Tocopherol (Vit.E)", "生育酚乙酸酯": "Tocopheryl Acetate",
    "神经酰胺 NP": "Ceramide NP", "神经酰胺 EOP": "Ceramide EOP",
    "神经酰胺 AP": "Ceramide AP", "神经酰胺 NS": "Ceramide NS",
    "神经酰胺NP": "Ceramide NP", "神经酰胺EOP": "Ceramide EOP",
    "角鲨烷": "Squalane", "角鲨烯": "Squalene",
    "泛醇": "Panthenol", "尿囊素": "Allantoin",
    "积雪草苷": "Asiaticoside", "积雪草酸": "Asiatic Acid",
    "羟基积雪草酸": "Madecassic Acid",
    "积雪草（CENTELLA ASIATICA）提取物": "Centella Asiatica Extract",
    "积雪草": "Centella Asiatica",
    "胶原": "Collagen", "水解胶原": "Hydrolyzed Collagen",
    "肌肽": "Carnosine", "谷胱甘肽": "Glutathione",
    "虾青素": "Astaxanthin", "辅酶Q10": "Coenzyme Q10",
    "麦角硫因": "Ergothioneine", "凝血酸": "Tranexamic Acid",
    # 保湿/调理
    "甜菜碱": "Betaine", "海藻糖": "Trehalose", "甘露糖醇": "Mannitol",
    "赤藓醇": "Erythritol", "木糖醇": "Xylitol", "山梨醇": "Sorbitol",
    "PEG-40氢化蓖麻油": "PEG-40 Hydrogenated Castor Oil",
    "卡波姆": "Carbomer", "丙烯酸钠": "Sodium Acrylate",
    "黄原胶": "Xanthan Gum", "卡拉胶": "Carrageenan",
    "羟乙基纤维素": "Hydroxyethylcellulose",
    "羟丙基甲基纤维素": "Hydroxypropyl Methylcellulose",
    # 防腐/稳定
    "苯氧乙醇": "Phenoxyethanol", "乙基己基甘油": "Ethylhexylglycerin",
    "1,2-己二醇": "1,2-Hexanediol", "辛酰羟肟酸": "Caprylyl Hydroxamic Acid",
    "对羟基苯乙酮": "p-Hydroxyacetophenone",
    "EDTA二钠": "Disodium EDTA", "EDTA": "EDTA",
    "苯甲酸钠": "Sodium Benzoate",
    # 防晒
    "甲氧基肉桂酸乙基己酯": "Ethylhexyl Methoxycinnamate",
    "二苯酮-3": "Benzophenone-3 (Oxybenzone)",
    "氧化锌": "Zinc Oxide", "二氧化钛": "Titanium Dioxide",
    # 乳化/表活
    "鲸蜡醇": "Cetyl Alcohol", "硬脂醇": "Stearyl Alcohol",
    "鲸蜡硬脂醇": "Cetearyl Alcohol",
    "聚二甲基硅氧烷": "Dimethicone", "环五聚二甲基硅氧烷": "Cyclopentasiloxane",
    "环甲硅油": "Cyclomethicone",
    "椰油酰两性基二乙酸二钠": "Disodium Cocoamphodiacetate",
    "月桂醇聚醚硫酸酯钠": "Sodium Laureth Sulfate",
    "烷基葡萄糖苷": "Alkyl Glucoside", "椰油葡糖苷": "Coco-Glucoside",
    # 植物提取
    "白池花（LIMNANTHES ALBA）籽油": "Meadowfoam Seed Oil",
    "霍霍巴（SIMMONDSIA CHINENSIS）籽油": "Jojoba Seed Oil",
    "向日葵（HELIANTHUS ANNUUS）籽油": "Sunflower Seed Oil",
    "山茶（CAMELLIA JAPONICA）籽油": "Camellia Japonica Seed Oil",
    "角鲨烷": "Squalane",
    "薰衣草（LAVANDULA ANGUSTIFOLIA）花提取物": "Lavender Flower Extract",
    "人参": "Panax Ginseng",
    # 其他常见
    "香精": "Fragrance (Parfum)", "香料": "Fragrance",
    "色素": "Pigment/Colorant",
    "氯化钠": "Sodium Chloride (Salt)",
    "柠檬酸": "Citric Acid", "苹果酸": "Malic Acid",
    "精氨酸": "Arginine", "赖氨酸": "Lysine",
}


def _translate_ingr(zh: str) -> str:
    """翻译单个成分名：精确匹配→局部匹配→原文"""
    key = zh.strip()
    if key in _INGR_ZH2EN:
        return _INGR_ZH2EN[key]
    # 局部匹配（取最长匹配的英文名）
    best = ""
    for zh_k, en_v in _INGR_ZH2EN.items():
        if zh_k in key and len(zh_k) > len(best):
            best = en_v
    # 如有括号内的INCI名，直接提取
    import re
    inci = re.findall(r"[A-Z][A-Z\-\s]+[A-Z]", key)
    if inci:
        return inci[0].title()
    return best


def _parse_ingr(s: str) -> list[str]:
    """解析成分列表（逗号/中文逗号分隔）"""
    import re
    items = re.split(r"[，,、；\[\]【】\n]", s or "")
    return [i.strip().strip("0123456789. ") for i in items if i.strip() and len(i.strip()) > 1]


# ─────────────────────────────────────────────
# 悬浮成分对比层：拖拽 SKU 卡片放入 → JS 端计算成分差异
# （整页月历 + 悬浮层渲染在同一个 components.v1.html 文档里，
#   这样拖拽事件才能在同一 DOM 内被捕获）
# ─────────────────────────────────────────────
_FLOATING_CMP_CSS = """
<style>
  .prod-card[draggable="true"] { cursor: grab; }
  .prod-card[draggable="true"]:active { cursor: grabbing; }
  .prod-card[draggable="true"] summary::before {
    content: "⠿"; color: #c5cfe0; font-size: 10px; margin-right: 4px;
  }

  #cmp-toggle-btn {
    position: fixed; right: 20px; top: 18px;
    background: #1565c0; color: #fff; border: none; border-radius: 30px;
    padding: 8px 16px; font-size: 13px; font-weight: 600; cursor: pointer;
    box-shadow: 0 4px 14px rgba(0,0,0,.25); z-index: 9999;
    display: flex; align-items: center; gap: 6px;
  }
  #cmp-toggle-btn:hover { background: #0d47a1; }

  #cmp-drawer {
    position: fixed; left: 0; right: 0; bottom: 0; height: 0; overflow: hidden;
    background: #fff; border-top: 2px solid #1565c0;
    box-shadow: 0 -6px 20px rgba(0,0,0,.18);
    transition: height .28s ease; z-index: 9998;
  }
  #cmp-drawer.open { height: 40vh; min-height: 120px; }
  /* 拖拽调高把手 */
  #cmp-resize-handle {
    position: absolute; top: 0; left: 0; right: 0; height: 10px;
    cursor: ns-resize; background: transparent;
    display: flex; justify-content: center; align-items: center; z-index: 10;
  }
  #cmp-resize-handle::after {
    content: '━━━━━━'; color: #c5cfe0; font-size: 8px; letter-spacing: 3px;
    pointer-events: none;
  }
  #cmp-resize-handle:hover::after { color: #1565c0; }
  #cmp-drawer-inner { padding: 14px 22px; height: 100%; box-sizing: border-box; overflow-y: auto; }
  #cmp-drawer-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px; }
  #cmp-drawer-header h3 { margin: 0; font-size: 15px; color: #1a2b4a; font-family: 'Inter', sans-serif; }
  .cmp-close-btn { background: #f0f2f5; border: none; font-size: 13px; cursor: pointer;
                   color: #5f7089; border-radius: 6px; padding: 3px 9px; margin-left: 6px; }
  .cmp-close-btn:hover { background: #e2e8f0; }

  .cmp-slots { display: flex; gap: 14px; margin-bottom: 10px; }
  .cmp-slot {
    flex: 1; min-height: 66px; border: 2px dashed #c5cfe0; border-radius: 8px;
    display: flex; align-items: center; justify-content: center; padding: 8px;
    font-size: 12px; color: #9aa5b1; text-align: center; transition: .15s;
  }
  .cmp-slot.filled { border-style: solid; border-color: #1565c0; background: #f5f9ff;
                     color: #1a2b4a; flex-direction: column; align-items: flex-start; text-align: left; }
  .cmp-slot.dragover { background: #e3f2fd; border-color: #1565c0; }
  .cmp-slot .slot-brand { font-weight: 700; font-size: 12px; }
  .cmp-slot .slot-name  { font-size: 11px; margin: 2px 0; }
  .cmp-slot .slot-reg   { font-size: 9px; color: #666; font-family: monospace; }
  .cmp-slot .slot-clear { align-self: flex-end; background: #eee; border: none; border-radius: 4px;
                           font-size: 10px; padding: 2px 6px; cursor: pointer; margin-top: 4px; }

  .cmp-grid { display: grid; grid-template-columns: 70px 1fr 60px 60px; gap: 3px; font-size: 12px; }
  .cmp-head { font-weight: 700; color: #8898aa; font-size: 10px; text-transform: uppercase;
              padding: 4px 6px; background: #f8f9fb; border-radius: 4px; }
  .cmp-row  { display: contents; }
  .cmp-cell { padding: 4px 6px; border-bottom: 1px solid #f5f5f5; }
  .cmp-cell.same    { color: #374151; }
  .cmp-cell.moved-u { color: #1565c0; background: #e3f2fd; border-radius: 3px; }
  .cmp-cell.moved-d { color: #e65100; background: #fff3e0; border-radius: 3px; }
  .cmp-cell.added   { color: #1b5e20; background: #e8f5e9; border-radius: 3px; }
  .cmp-cell.removed { color: #b71c1c; background: #ffebee; border-radius: 3px; text-decoration: line-through; }

  #print-btn {
    position: fixed; right: 150px; top: 18px;
    background: #37474f; color: #fff; border: none; border-radius: 30px;
    padding: 8px 16px; font-size: 13px; font-weight: 600; cursor: pointer;
    box-shadow: 0 4px 14px rgba(0,0,0,.25); z-index: 9999;
    display: flex; align-items: center; gap: 6px;
  }
  #print-btn:hover { background: #1a2b4a; }

  /* ── 打印样式：让整份月历(所有品牌/季度)完整分页打印，
     成分对比抽屉强制换页放到最后一页 ── */
  @media print {
    @page { size: A4 landscape; margin: 10mm; }
    #print-btn, #cmp-toggle-btn, #cmp-resize-handle, .cmp-close-btn { display: none !important; }
    body { background: #fff !important; padding-bottom: 0 !important; }
    .cal-wrap { box-shadow: none !important; }
    .cal-th-brand, .cal-brand { position: static !important; }
    .cal-table tr { page-break-inside: avoid; }
    #cmp-drawer {
      position: static !important;
      height: auto !important;
      min-height: 0 !important;
      overflow: visible !important;
      box-shadow: none !important;
      page-break-before: always;
    }
    #cmp-drawer-inner { overflow: visible !important; height: auto !important; }
  }
</style>
"""


def _build_products_map(filtered: list[dict]) -> str:
    """构建供悬浮层 JS 使用的产品字典（含已翻译成分列表），返回 JSON 字符串"""
    products: dict[str, dict] = {}
    for r in filtered:
        pid = r.get("_pid")
        if not pid:
            continue
        zh_list = _parse_ingr(r.get("ingredients", ""))
        en_list = [(_translate_ingr(x) or "") for x in zh_list]
        # 每条成分：用英文名做匹配键，同时携带中文原名用于双语显示
        ingr_items = [
            {"key": (en or zh).lower(), "label": (f"{zh} / {en}" if en and en != zh else zh)}
            for zh, en in zip(zh_list, en_list)
        ]
        products[pid] = {
            "name": r["name"],
            "brand": r["brand_en"],
            "reg": r["reg_num"],
            "date": str(r["notif_date"]),
            "ingr": ingr_items,
        }
    return json.dumps(products, ensure_ascii=False)


def _build_compare_widget_html(products_json: str) -> str:
    """悬浮开关按钮 + 可展开抽屉（两个拖拽槽 + JS 成分对比表）"""
    return f"""
<button id="print-btn" onclick="window.print()" title="打印/导出PDF：会完整打印所有品牌与季度，成分对比放在最后一页">🖨️ Print</button>
<button id="cmp-toggle-btn" onclick="cmpToggleDrawer()">🔬 Compare <span id="cmp-badge">0/2</span></button>
<div id="cmp-drawer">
  <div id="cmp-resize-handle" title="Drag to resize"></div>
  <div id="cmp-drawer-inner">
    <div id="cmp-drawer-header">
      <h3>🔬 Ingredient Comparison — drag 2 SKU cards here</h3>
      <div>
        <button class="cmp-close-btn" title="Clear both" onclick="cmpClearAll()">🗑 Clear</button>
        <button class="cmp-close-btn" title="Close" onclick="cmpToggleDrawer(false)">✕ Close</button>
      </div>
    </div>
    <div class="cmp-slots">
      <div class="cmp-slot" id="cmp-slot-A"
           ondragover="cmpAllowDrop(event)" ondragleave="cmpDragLeave(event)" ondrop="cmpDropSlot(event,'A')">
        Drag SKU A here
      </div>
      <div class="cmp-slot" id="cmp-slot-B"
           ondragover="cmpAllowDrop(event)" ondragleave="cmpDragLeave(event)" ondrop="cmpDropSlot(event,'B')">
        Drag SKU B here
      </div>
    </div>
    <div id="cmp-diff-area">
      <p style="color:#9aa5b1;font-size:12px;margin-top:6px">
        Drop two SKU cards above to compare ingredients (🟢 added · 🔴 removed · 🔵 moved up ≥5 · 🟠 moved down ≥5).
      </p>
    </div>
  </div>
</div>
<script>
  var PRODUCTS = {products_json};
  var cmpSlots = {{A: null, B: null}};

  function cmpDragStart(ev) {{
    var pid = ev.currentTarget.getAttribute('data-pid');
    if (!pid) {{ ev.preventDefault(); return; }}
    ev.dataTransfer.setData('text/plain', pid);
    ev.dataTransfer.effectAllowed = 'copy';
  }}
  function cmpAllowDrop(ev) {{ ev.preventDefault(); ev.currentTarget.classList.add('dragover'); }}
  function cmpDragLeave(ev) {{ ev.currentTarget.classList.remove('dragover'); }}
  function cmpDropSlot(ev, key) {{
    ev.preventDefault();
    ev.currentTarget.classList.remove('dragover');
    var pid = ev.dataTransfer.getData('text/plain');
    if (!pid || !(pid in PRODUCTS)) return;
    cmpSlots[key] = pid;
    cmpRenderSlots(); cmpRenderDiff(); cmpUpdateBadge();
    var drawer = document.getElementById('cmp-drawer');
    if (!drawer.classList.contains('open')) drawer.classList.add('open');
  }}
  function cmpClearSlot(key) {{
    cmpSlots[key] = null;
    cmpRenderSlots(); cmpRenderDiff(); cmpUpdateBadge();
  }}
  function cmpClearAll() {{
    cmpSlots = {{A: null, B: null}};
    cmpRenderSlots(); cmpRenderDiff(); cmpUpdateBadge();
  }}
  function cmpToggleDrawer(force) {{
    var d = document.getElementById('cmp-drawer');
    if (force === true) d.classList.add('open');
    else if (force === false) d.classList.remove('open');
    else d.classList.toggle('open');
  }}
  function cmpUpdateBadge() {{
    var n = (cmpSlots.A ? 1 : 0) + (cmpSlots.B ? 1 : 0);
    document.getElementById('cmp-badge').innerText = n + '/2';
  }}
  function cmpRenderSlots() {{
    ['A', 'B'].forEach(function(key) {{
      var el = document.getElementById('cmp-slot-' + key);
      var pid = cmpSlots[key];
      if (!pid) {{
        el.className = 'cmp-slot';
        el.innerHTML = 'Drag SKU ' + key + ' here';
      }} else {{
        var p = PRODUCTS[pid];
        el.className = 'cmp-slot filled';
        el.innerHTML = '<div class="slot-brand">' + key + ' · ' + p.brand + '</div>' +
          '<div class="slot-name">' + p.name + '</div>' +
          '<div class="slot-reg">' + (p.reg || '') + '</div>' +
          '<button class="slot-clear" onclick="cmpClearSlot(\\'' + key + '\\')">✕ remove</button>';
      }}
    }});
  }}
  function cmpIngrDiff(listA, listB) {{
    // listA/listB 现在是 {{key, label}} 对象数组
    var posA = {{}}, posB = {{}};
    listA.forEach(function(x, i) {{ if (!(x.key in posA)) posA[x.key] = i; }});
    listB.forEach(function(x, i) {{ if (!(x.key in posB)) posB[x.key] = i; }});
    var rows = [];
    listA.forEach(function(item, i) {{
      if (item.key in posB) {{
        var j = posB[item.key], shift = i - j;
        rows.push({{label: item.label, status: Math.abs(shift) >= 5 ? 'moved' : 'same',
                    posA: i + 1, posB: j + 1, shift: shift}});
      }} else {{
        rows.push({{label: item.label, status: 'removed', posA: i + 1, posB: null, shift: null}});
      }}
    }});
    listB.forEach(function(item, j) {{
      if (!(item.key in posA)) rows.push({{label: item.label, status: 'added', posA: null, posB: j + 1, shift: null}});
    }});
    rows.sort(function(a, b) {{
      var sa = a.posB !== null ? a.posB : (a.posA || 999) + 5000;
      var sb = b.posB !== null ? b.posB : (b.posA || 999) + 5000;
      return sa - sb;
    }});
    return rows;
  }}
  function cmpRenderDiff() {{
    var container = document.getElementById('cmp-diff-area');
    var a = cmpSlots.A ? PRODUCTS[cmpSlots.A] : null;
    var b = cmpSlots.B ? PRODUCTS[cmpSlots.B] : null;
    if (!a || !b) {{
      container.innerHTML = '<p style="color:#9aa5b1;font-size:12px;margin-top:6px">' +
        'Drop two SKU cards above to compare ingredients (🟢 added · 🔴 removed · 🔵 moved up ≥5 · 🟠 moved down ≥5).</p>';
      return;
    }}
    if (!a.ingr.length || !b.ingr.length) {{
      container.innerHTML = '<p style="color:#e65100;font-size:12px;margin-top:6px">' +
        '⚠️ One of the selected products has no ingredient data.</p>';
      return;
    }}
    var diff = cmpIngrDiff(a.ingr, b.ingr);
    var added = diff.filter(function(d) {{ return d.status === 'added'; }}).length;
    var removed = diff.filter(function(d) {{ return d.status === 'removed'; }}).length;
    var up = diff.filter(function(d) {{ return d.status === 'moved' && d.shift > 0; }}).length;
    var down = diff.filter(function(d) {{ return d.status === 'moved' && d.shift < 0; }}).length;

    var rowsHtml = '';
    diff.forEach(function(d) {{
      var cls, arrow, pa, pb;
      if (d.status === 'added') {{ cls = 'added'; arrow = '🟢 NEW'; pa = '—'; pb = d.posB; }}
      else if (d.status === 'removed') {{ cls = 'removed'; arrow = '🔴 OUT'; pa = d.posA; pb = '—'; }}
      else if (d.status === 'moved') {{
        if (d.shift > 0) {{ cls = 'moved-u'; arrow = '🔵 ↑' + d.shift; }}
        else {{ cls = 'moved-d'; arrow = '🟠 ↓' + Math.abs(d.shift); }}
        pa = d.posA; pb = d.posB;
      }} else {{ cls = 'same'; arrow = '='; pa = d.posA; pb = d.posB; }}
      rowsHtml += '<div class="cmp-row">' +
        '<div class="cmp-cell ' + cls + '" style="text-align:center">' + arrow + '</div>' +
        '<div class="cmp-cell ' + cls + '">' + d.label + '</div>' +
        '<div class="cmp-cell ' + cls + '" style="text-align:center">' + pa + '</div>' +
        '<div class="cmp-cell ' + cls + '" style="text-align:center">' + pb + '</div>' +
        '</div>';
    }});

    container.innerHTML =
      '<div style="font-size:12px;margin:6px 0 8px">' +
      '<span style="color:#1b5e20"><b>+' + added + ' Added</b></span> &nbsp; ' +
      '<span style="color:#b71c1c"><b>-' + removed + ' Removed</b></span> &nbsp; ' +
      '<span style="color:#1565c0"><b>↑' + up + ' Moved Up ≥5</b></span> &nbsp; ' +
      '<span style="color:#e65100"><b>↓' + down + ' Moved Down ≥5</b></span>' +
      '</div>' +
      '<div class="cmp-grid">' +
      '<div class="cmp-head">Change</div><div class="cmp-head">Ingredient</div>' +
      '<div class="cmp-head">Pos A</div><div class="cmp-head">Pos B</div>' +
      rowsHtml + '</div>';
  }}

  cmpRenderSlots(); cmpUpdateBadge();

  // 拖拽调整抽屉高度
  var _cmpDragY = 0, _cmpDragH = 0, _cmpDragging = false;
  document.getElementById('cmp-resize-handle').addEventListener('mousedown', function(e) {{
    var drawer = document.getElementById('cmp-drawer');
    if (!drawer.classList.contains('open')) return;
    _cmpDragging = true;
    _cmpDragY = e.clientY;
    _cmpDragH = drawer.offsetHeight;
    drawer.style.transition = 'none';
    e.preventDefault();
  }});
  document.addEventListener('mousemove', function(e) {{
    if (!_cmpDragging) return;
    var delta = _cmpDragY - e.clientY;
    var newH = Math.max(120, Math.min(window.innerHeight * 0.9, _cmpDragH + delta));
    document.getElementById('cmp-drawer').style.height = newH + 'px';
  }});
  document.addEventListener('mouseup', function() {{
    if (_cmpDragging) {{
      _cmpDragging = false;
      document.getElementById('cmp-drawer').style.transition = '';
    }}
  }});
</script>
"""


def _render_pdf_download_section(records: list[dict]):
    """
    扫描本地已下载的 PDF 文件（DOWNLOAD_BASE/{brand}/{name}.pdf 或 特化--{name}.pdf），
    提供 Streamlit 原生 download_button 一键下载。
    也列出 Artwork (H列) 和 mini-POC (J列) 链接供直接访问。
    """
    base = Path(DOWNLOAD_BASE)
    if not base.exists():
        return

    # 收集本地已存在的 PDF（按品牌→产品名建立索引）
    local_pdfs: dict[str, Path] = {}  # key = "{brand_en}:{prod_name}" → Path
    for brand_en in BRANDS:
        folder = base / brand_en
        if not folder.exists():
            continue
        for pdf in folder.glob("*.pdf"):
            # 兼容 "特化--{name}.pdf" 和 "{name}.pdf"
            stem = pdf.stem
            if stem.startswith("特化--"):
                stem = stem[3:]
            local_pdfs[f"{brand_en}:{stem}"] = pdf

    if not local_pdfs and not any(r.get("pdf_url") or r.get("poc_url") for r in records):
        return

    with st.expander(f"⬇️ PDF 本地下载（{len(local_pdfs)} 个文件）— 公司内网内使用此处下载", expanded=False):
        # 按品牌分组显示
        by_brand: dict[str, list] = {}
        for r in records:
            by_brand.setdefault(r["brand_en"], []).append(r)

        for brand_en, prods in by_brand.items():
            # 只显示有链接或有本地文件的品牌
            brand_entries = []
            for p in prods:
                key_exact  = f"{brand_en}:{p['name']}"
                key_prefix = next(
                    (k for k in local_pdfs if k.startswith(f"{brand_en}:") and p["name"][:15] in k),
                    None
                )
                local_path = local_pdfs.get(key_exact) or (local_pdfs.get(key_prefix) if key_prefix else None)
                has_link = p.get("pdf_url") or p.get("poc_url") or local_path
                if has_link:
                    brand_entries.append((p, local_path))

            if not brand_entries:
                continue

            st.markdown(f"**{brand_en} / {BRANDS[brand_en]}** ({len(brand_entries)}件)")
            for i, (p, local_path) in enumerate(brand_entries):
                cols = st.columns([4, 1, 1, 1])
                cols[0].markdown(
                    f"<span style='font-size:12px'>{p['name'][:40]}</span>",
                    unsafe_allow_html=True
                )
                # 本地 PDF 下载
                if local_path:
                    pdf_bytes = local_path.read_bytes()
                    cols[1].download_button(
                        "⬇ PDF",
                        data=pdf_bytes,
                        file_name=local_path.name,
                        mime="application/pdf",
                        key=f"dl_{brand_en}_{p['name'][:20]}_{i}",
                        use_container_width=True
                    )
                else:
                    cols[1].markdown("<span style='color:#aaa;font-size:11px'>无本地文件</span>",
                                     unsafe_allow_html=True)
                # Artwork 链接（NMPA PDF viewer）
                if p.get("pdf_url"):
                    cols[2].markdown(
                        f"<a href='{p['pdf_url']}' target='_blank' "
                        f"style='font-size:11px;color:#1565c0'>🖼 Artwork</a>",
                        unsafe_allow_html=True
                    )
                # mini-POC 链接
                if p.get("poc_url"):
                    cols[3].markdown(
                        f"<a href='{p['poc_url']}' target='_blank' "
                        f"style='font-size:11px;color:#2e7d32'>🧪 POC</a>",
                        unsafe_allow_html=True
                    )


# ─────────────────────────────────────────────
# 管理模式：登录 / 触发抓取 / 增删改 + 操作日志
# ─────────────────────────────────────────────
ADMIN_LOG_FILE = LOG_DIR / "admin_actions.log"
PIPELINE_LOCK  = LOG_DIR / "pipeline.lock"

_EDITABLE_COLS = [
    ("upload_date", COL_UPLOAD_DATE, "上传日期 (MM/DD/YYYY)"),
    ("name",        COL_NAME,        "产品名称"),
    ("effect",      COL_EFFECT,      "功效宣称"),
    ("notif_date",  COL_DATE,        "备案/通知日期"),
    ("reg_num",     COL_REG_NUM,     "备案/注册号"),
    ("category",    COL_CATEGORY,    "类目/其他"),
    ("ingredients", COL_INGREDIENTS, "成分列表"),
    ("pdf_url",     COL_PDF_URL,     "PDF链接"),
    ("label_url",   COL_LABEL_URL,   "标签链接(Artwork)"),
    ("poc_url",     COL_POC_URL,     "mini POC链接"),
]


def log_admin_action(username: str, action: str, brand: str, detail: dict):
    ADMIN_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    entry = {
        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user": username, "action": action, "brand": brand, "detail": detail,
    }
    with open(ADMIN_LOG_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


def is_pipeline_running() -> bool:
    return PIPELINE_LOCK.exists()


def trigger_pipeline(days: int = 28):
    """后台（非阻塞）启动一次完整抓取流程，跑完后自动清理锁文件"""
    PIPELINE_LOCK.parent.mkdir(parents=True, exist_ok=True)
    PIPELINE_LOCK.write_text(datetime.now().isoformat(), encoding="utf-8")
    main_py = Path(__file__).parent / "main.py"
    python_exe = sys.executable
    if os.name == "nt":
        cmd = f'"{python_exe}" "{main_py}" --days {days} & del /f /q "{PIPELINE_LOCK}"'
        subprocess.Popen(["cmd", "/c", cmd], cwd=str(main_py.parent),
                          creationflags=subprocess.CREATE_NO_WINDOW)
    else:
        cmd = f'"{python_exe}" "{main_py}" --days {days}; rm -f "{PIPELINE_LOCK}"'
        subprocess.Popen(["/bin/sh", "-c", cmd], cwd=str(main_py.parent))


def get_last_run_info():
    if not LOG_DIR.exists():
        return None
    logs = sorted(LOG_DIR.glob("ci_bot_*.log"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not logs:
        return None
    latest = logs[0]
    try:
        tail = latest.read_text(encoding="utf-8", errors="ignore").splitlines()[-15:]
    except Exception:
        tail = []
    ok = any(("✅" in l and ("邮件" in l or "完成" in l)) for l in tail)
    return {
        "file": latest.name,
        "mtime": datetime.fromtimestamp(latest.stat().st_mtime),
        "tail": tail,
        "ok": ok,
    }


def render_login_box():
    """侧边栏：管理员登录 / 已登录状态显示"""
    with st.sidebar:
        st.markdown("### 🔒 管理模式")
        if st.session_state.get("is_admin"):
            st.success(f"已登录：{st.session_state.get('admin_user')}")
            if st.button("退出登录"):
                st.session_state["is_admin"] = False
                st.session_state["admin_user"] = None
                st.rerun()
        else:
            with st.form("admin_login_form", clear_on_submit=True):
                u = st.text_input("用户名")
                p = st.text_input("密码", type="password")
                submitted = st.form_submit_button("登录")
                if submitted:
                    if verify_login(u, p):
                        st.session_state["is_admin"] = True
                        st.session_state["admin_user"] = u
                        st.rerun()
                    else:
                        st.error("用户名或密码错误")


def render_admin_panel(records: list[dict]):
    """管理员面板：触发抓取 + 逐行增/删/改（仅作用于当前 EXCEL_PATH 文件）"""
    username = st.session_state.get("admin_user", "unknown")

    with st.expander("🛠️ 管理面板", expanded=False):
        # ── 抓取状态 & 手动触发 ──
        st.markdown("#### 📡 抓取任务")
        info = get_last_run_info()
        if info:
            status_icon = "✅" if info["ok"] else "⚠️"
            st.caption(f"{status_icon} 最近一次日志：{info['file']}（{info['mtime'].strftime('%Y-%m-%d %H:%M:%S')}）")
            with st.expander("查看最近日志尾部"):
                st.code("\n".join(info["tail"]))
        else:
            st.caption("暂无历史抓取日志")

        if is_pipeline_running():
            st.info("⏳ 有一次抓取任务正在后台运行中，请稍后刷新查看结果")
        else:
            days = st.number_input("查询天数", min_value=1, max_value=730, value=28, step=1)
            if st.button("🚀 立即触发一次全量抓取（后台运行，不会卡住页面）"):
                trigger_pipeline(days=int(days))
                log_admin_action(username, "trigger_pipeline", "-", {"days": int(days)})
                st.success("已在后台启动，可稍后刷新本页查看'最近一次日志'确认结果")
                st.rerun()

        st.divider()

        # ── 逐行增 / 删 / 改（仅当前 EXCEL_PATH 文件可编辑） ──
        st.markdown("#### ✏️ 数据增/删/改（仅影响当前主文件，历史归档只读）")
        editable_brands = sorted({r["brand_en"] for r in records if r.get("editable")})
        if not editable_brands:
            st.caption("当前主文件里没有可编辑的品牌数据")
            return

        brand_en = st.selectbox("选择品牌", editable_brands,
                                 format_func=lambda k: f"{k} / {BRANDS.get(k, '')}")
        brand_rows = sorted(
            [r for r in records if r.get("editable") and r["brand_en"] == brand_en],
            key=lambda r: r["row_idx"],
        )

        import pandas as pd
        preview_df = pd.DataFrame([
            {"行号": r["row_idx"], "产品名称": r["name"], "备案/通知日期": str(r["notif_date"]),
             "备案/注册号": r["reg_num"]}
            for r in brand_rows
        ])
        st.dataframe(preview_df, use_container_width=True, height=250)

        action = st.radio("操作", ["编辑已有行", "删除已有行", "新增一行"], horizontal=True)

        if action in ("编辑已有行", "删除已有行") and brand_rows:
            row_options = {r["row_idx"]: f"行{r['row_idx']} · {r['name']}" for r in brand_rows}
            sel_row_idx = st.selectbox("选择行", list(row_options.keys()),
                                        format_func=lambda i: row_options[i])
            sel_row = next(r for r in brand_rows if r["row_idx"] == sel_row_idx)

            if action == "编辑已有行":
                with st.form(f"edit_form_{brand_en}_{sel_row_idx}"):
                    new_vals = {}
                    for key, _col, label in _EDITABLE_COLS:
                        new_vals[key] = st.text_input(label, value=str(sel_row.get(key, "") or ""))
                    confirm = st.checkbox("✅ 确认保存以上修改")
                    if st.form_submit_button("保存修改") and confirm:
                        _write_row(EXCEL_PATH, brand_en, sel_row_idx, new_vals)
                        log_admin_action(username, "edit_row", brand_en,
                                          {"row_idx": sel_row_idx, "before": {k: sel_row.get(k) for k, *_ in _EDITABLE_COLS},
                                           "after": new_vals})
                        st.cache_data.clear()
                        st.success(f"已保存第 {sel_row_idx} 行的修改")
                        st.rerun()

            elif action == "删除已有行":
                st.warning(f"即将删除：行{sel_row_idx} · {sel_row['name']}（{sel_row['reg_num']}）")
                confirm_del = st.checkbox("✅ 我确认要删除这一行（不可撤销，请谨慎操作）")
                if st.button("🗑 确认删除此行", disabled=not confirm_del):
                    _delete_row(EXCEL_PATH, brand_en, sel_row_idx)
                    log_admin_action(username, "delete_row", brand_en,
                                      {"row_idx": sel_row_idx, "deleted": {k: sel_row.get(k) for k, *_ in _EDITABLE_COLS}})
                    st.cache_data.clear()
                    st.success(f"已删除第 {sel_row_idx} 行")
                    st.rerun()

        elif action == "新增一行":
            with st.form(f"add_form_{brand_en}"):
                new_vals = {}
                for key, _col, label in _EDITABLE_COLS:
                    default = datetime.now().strftime("%m/%d/%Y") if key == "upload_date" else ""
                    new_vals[key] = st.text_input(label, value=default)
                confirm_add = st.checkbox("✅ 确认新增这一行")
                if st.form_submit_button("新增") and confirm_add:
                    new_row_idx = _append_row(EXCEL_PATH, brand_en, new_vals)
                    log_admin_action(username, "add_row", brand_en, {"row_idx": new_row_idx, "values": new_vals})
                    st.cache_data.clear()
                    st.success(f"已新增第 {new_row_idx} 行")
                    st.rerun()


def _backup_excel(path, keep_last: int = 30) -> Path:
    """写入前自动备份 Excel，供 _write_row/_delete_row/_append_row 共用。
    误删/误改可从 <文件所在目录>/_admin_backups/ 里恢复最近的一份。
    只保留最近 keep_last 份，避免频繁编辑时备份无限堆积。"""
    src = Path(path)
    backup_dir = src.parent / "_admin_backups"
    backup_dir.mkdir(exist_ok=True)
    backup_path = backup_dir / f"{src.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy2(src, backup_path)

    old_backups = sorted(backup_dir.glob(f"{src.stem}_backup_*.xlsx"), key=lambda p: p.stat().st_mtime)
    for stale in old_backups[:-keep_last]:
        stale.unlink(missing_ok=True)

    return backup_path


def _write_row(path, brand_en: str, row_idx: int, vals: dict):
    _backup_excel(path)
    wb = openpyxl.load_workbook(path)
    ws = wb[brand_en]
    for key, col, _label in _EDITABLE_COLS:
        ws.cell(row=row_idx, column=col, value=vals.get(key, ""))
    wb.save(path)
    wb.close()


def _delete_row(path, brand_en: str, row_idx: int):
    _backup_excel(path)
    wb = openpyxl.load_workbook(path)
    ws = wb[brand_en]
    ws.delete_rows(row_idx, 1)
    wb.save(path)
    wb.close()


def _append_row(path, brand_en: str, vals: dict) -> int:
    _backup_excel(path)
    wb = openpyxl.load_workbook(path)
    ws = wb[brand_en]
    new_row_idx = ws.max_row + 1
    for key, col, _label in _EDITABLE_COLS:
        ws.cell(row=new_row_idx, column=col, value=vals.get(key, ""))
    wb.save(path)
    wb.close()
    return new_row_idx


# ─────────────────────────────────────────────
# 主界面
# ─────────────────────────────────────────────
def main():
    st.markdown(GLOBAL_CSS, unsafe_allow_html=True)

    st.markdown("""
    <div class="hero">
      <h1>🔬 Competitor Intelligence · New SKU Tracker</h1>
      <p>竞品注册 & 备案月历 · 成分 · 功效宣称 · Artwork PDF · 功效证明</p>
    </div>
    """, unsafe_allow_html=True)

    render_login_box()

    records = load_data()
    if not records:
        st.error("⚠️ 未读到数据，请先运行：`python src/main.py --days 730`")
        st.stop()

    if st.session_state.get("is_admin"):
        render_admin_panel(records)

    img_map = load_image_map()

    all_brand_keys = list(BRANDS.keys())
    all_years = sorted({r["year"] for r in records}, reverse=True)

    # 品牌显示标签："PROYA / 珀莱雅"
    brand_labels = [f"{k} / {v}" for k, v in BRANDS.items()]
    label_to_key = {f"{k} / {v}": k for k, v in BRANDS.items()}
    default_labels = brand_labels  # 默认全选

    # ── 筛选栏 ──
    c1, c2, c3, c4, c5 = st.columns([3, 2, 1.5, 0.5, 0.5])

    with c1:
        sel_brand_labels = st.multiselect(
            "品牌", brand_labels, default=default_labels,
            label_visibility="collapsed",
            placeholder="Select brands (all by default)…",
        )
        selected_brands = [label_to_key[l] for l in sel_brand_labels if l in label_to_key]
    with c2:
        selected_years = st.multiselect(
            "Year (max 3)", all_years,
            default=all_years[:min(3, len(all_years))],
            max_selections=3,
            label_visibility="collapsed",
            placeholder="Select up to 3 years…",
        )
    with c3:
        reg_type_filter = st.selectbox(
            "Type", ["全部 All", "普通备案 Filing", "特殊注册 Registration"],
            label_visibility="collapsed",
        )
    with c4:
        if st.button("✔ All", use_container_width=True, help="Select all brands"):
            st.session_state["展开品牌"] = True  # trigger rerun with all selected
    with c5:
        show_en = st.toggle("EN", value=False, help="Show English product name translation")

    if not selected_brands:
        selected_brands = all_brand_keys
    if not selected_years:
        selected_years = all_years[:1]

    # ── 过滤 ──
    def _type_match(r):
        if "特殊" in reg_type_filter: return r["reg_type"] == "特殊注册"
        if "备案" in reg_type_filter: return r["reg_type"] == "普通备案"
        return True

    filtered = [r for r in records
                if r["brand_en"] in selected_brands
                and r["year"] in selected_years
                and _type_match(r)]

    # ── 统计摘要 ──
    total   = len(filtered)
    normal  = sum(1 for r in filtered if r["reg_type"] == "普通备案")
    special = sum(1 for r in filtered if r["reg_type"] == "特殊注册")
    brands_n = len({r["brand_en"] for r in filtered})

    st.markdown(f"""
    <div class="stat-row">
      <div class="stat-card blue"><div class="label">Total Products</div><div class="value">{total}</div></div>
      <div class="stat-card green"><div class="label">🟢 普通备案</div><div class="value">{normal}</div></div>
      <div class="stat-card red"><div class="label">🔴 特殊注册</div><div class="value">{special}</div></div>
      <div class="stat-card"><div class="label">Active Brands</div><div class="value">{brands_n}</div></div>
    </div>
    """, unsafe_allow_html=True)

    # ── 月历（每个选中年份一张）+ 悬浮成分对比层 ──
    # 翻译缓存（只在 EN 开关打开时填充）
    trans_cache: dict = {}
    if show_en:
        names = [r["name"] for r in filtered]
        trans_cache = translate_batch(names, get_translation_cache())

    # 悬浮层需要和卡片在同一个 DOM 内才能捕获拖拽事件，所以把所有年份的
    # 月历 + 悬浮开关按钮 + 抽屉一起放进一个 components.v1.html 文档里渲染。
    for i, r in enumerate(filtered):
        r["_pid"] = f"p{i}"

    quarter_labels = [("Q1", "JFM"), ("Q2", "AMJ"), ("Q3", "JAS"), ("Q4", "OND")]

    year_sections = ""
    for yr in sorted(selected_years, reverse=True):
        yr_filtered = [r for r in filtered if r["year"] == yr]
        if not yr_filtered:
            continue
        months = [(f"{yr}-{qk}", label) for qk, label in quarter_labels]
        cal_html = build_calendar_html(yr_filtered, selected_brands, months, trans_cache, img_map)
        year_sections += f"""
        <h3 style="color:#1a2b4a;margin:18px 0 6px;font-family:'Inter',sans-serif">📅 {yr}</h3>
        {cal_html}"""

    products_json = _build_products_map(filtered)
    compare_widget_html = _build_compare_widget_html(products_json)

    full_html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
{GLOBAL_CSS}
{_FLOATING_CMP_CSS}
</head><body style="margin:0;padding:8px 4px 70px;background:#f7f9fc">
{year_sections}
{compare_widget_html}
</body></html>"""

    # 固定高度 + 内部滚动条：悬浮按钮/抽屉的 position:fixed 是相对这个
    # iframe 自身可视区域的，只有高度固定、内容用内部滚动条滚动，才能让
    # 悬浮层在"滚动年份/品牌"时始终保持可见。
    FRAME_HEIGHT = 900
    st.iframe(full_html, height=FRAME_HEIGHT)

    # ── PDF 本地下载区（公司网络内 NMPA 外部链接可能无法打开，优先用此）──
    _render_pdf_download_section(filtered)

    # ── 原始数据 ──
    with st.expander("📊 原始数据"):
        import pandas as pd
        df = pd.DataFrame(filtered).drop(columns=["year", "month", "source_file", "_pid",
                                                  "editable", "row_idx"],
                                         errors="ignore")
        st.dataframe(df, use_container_width=True)


if __name__ == "__main__":
    main()
