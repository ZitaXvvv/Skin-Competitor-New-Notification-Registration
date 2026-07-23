"""
生成"730天历史抓取"里恢复出的68个产品的参考清单（不含日期，不接入看板日历）。
- 从已下载但保留在磁盘上的68个PDF中提取：品牌 / 真实产品名 / (特化类)备案号 / 全成分
- 用 fitz 把PDF首页渲染成缩略图，方便人工核对是哪款产品
- 汇总写入一个独立的 Excel 参考文件（文件名不匹配 dashboard 的 "CI_List_Ada *.xlsx" 通配符，
  不会被自动加载进看板日历——因为这些记录没有可靠日期，强行注入日期会误导日历视图）
"""
import re
import sys
from pathlib import Path
from datetime import datetime

import fitz
import openpyxl
from openpyxl.drawing.image import Image as XLImage

sys.path.insert(0, str(Path(__file__).parent))

BASE = Path(r"C:\Users\xie.x.3\Documents\Olay CI")
LO = datetime(2026, 7, 22, 16, 47, 0)
HI = datetime(2026, 7, 22, 17, 0, 0)

REG_PAT = re.compile(r"(国妆特字|国妆特进字|卫妆特字|妆网备字|国妆备字|国妆备进字)[0-9A-Za-z（）()]*")
OUT_DIR = BASE / "_Recovered_730d_Reference"
OUT_XLSX = BASE / "Recovered_730d_Reference.xlsx"


def extract_special_info(pdf_path: Path):
    doc = fitz.open(str(pdf_path))
    full = "".join(p.get_text() for p in doc)
    m = REG_PAT.search(full)
    reg_num = m.group(0) if m else None
    ing_m = re.search(r"全成分[:：]?\s*(.{0,500}?)(?:其他微量成分|注册人|生产企业|$)", full, re.S)
    ingredients = ing_m.group(1).strip().replace("\n", " ") if ing_m else ""
    doc.close()
    return reg_num, ingredients


def render_thumbnail(pdf_path: Path, out_png: Path, max_dim_px: int = 500):
    doc = fitz.open(str(pdf_path))
    page = doc[0]
    rect = page.rect
    scale = max_dim_px / max(rect.width, rect.height)
    mat = fitz.Matrix(scale, scale)
    pix = page.get_pixmap(matrix=mat)
    out_png.parent.mkdir(parents=True, exist_ok=True)
    pix.save(str(out_png))
    doc.close()


def main():
    records = []
    for d in sorted(BASE.iterdir()):
        if not d.is_dir() or d.name.startswith("_"):
            continue
        for f in sorted(d.glob("*.pdf")):
            mt = datetime.fromtimestamp(f.stat().st_mtime)
            if not (LO <= mt <= HI):
                continue
            is_special = f.name.startswith("特化--")
            name = f.name[len("特化--"):] if is_special else f.name
            name = name[:-4]
            rec = {"brand": d.name, "name": name, "is_special": is_special, "pdf": f}
            if is_special:
                reg_num, ingredients = extract_special_info(f)
                rec["reg_num"] = reg_num
                rec["ingredients"] = ingredients
            else:
                rec["reg_num"] = None
                rec["ingredients"] = ""
            records.append(rec)

    print(f"共 {len(records)} 条记录")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recovered"
    headers = ["品牌", "产品名称", "类型", "备案/注册号（仅特化类已知）", "全成分（仅特化类）",
               "原PDF路径", "缩略图"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 22
    ws.column_dimensions["G"].width = 20

    for i, rec in enumerate(records, start=2):
        thumb_png = OUT_DIR / rec["brand"] / f"{rec['name']}.png"
        try:
            render_thumbnail(rec["pdf"], thumb_png)
        except Exception as exc:
            print(f"  渲染失败 {rec['pdf'].name}: {exc}")
            thumb_png = None

        ws.cell(row=i, column=1, value=rec["brand"])
        ws.cell(row=i, column=2, value=rec["name"])
        ws.cell(row=i, column=3, value="特化(注册)" if rec["is_special"] else "普通备案")
        ws.cell(row=i, column=4, value=rec["reg_num"] or "")
        ws.cell(row=i, column=5, value=rec["ingredients"])
        ws.cell(row=i, column=6, value=str(rec["pdf"]))
        ws.row_dimensions[i].height = 90

        if thumb_png and thumb_png.exists():
            try:
                img = XLImage(str(thumb_png))
                img.height = 110
                img.width = 110
                ws.add_image(img, f"G{i}")
            except Exception as exc:
                print(f"  插入缩略图失败 {thumb_png}: {exc}")

    wb.save(OUT_XLSX)
    print(f"✅ 已生成参考文件: {OUT_XLSX}")
    print(f"   缩略图目录: {OUT_DIR}")


if __name__ == "__main__":
    main()
