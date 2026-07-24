"""
Hero SKU Cross Brand Matrix
============================
列：护肤步骤（Cleanser → Other，共15列）
行：品牌 × 系列
格：该品牌-系列在该护肤步骤上的明星单品卡片

品牌/系列数据来自：Skincare Competitor Product Menu Mar'26.docx 第5页（Appendix）
"""
import sys
from pathlib import Path
import streamlit as st

sys.path.insert(0, str(Path(__file__).parent.parent))

# ──────────────────────────────────────────────
# 页面配置
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="Hero SKU Matrix",
    page_icon="🏆",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────────
# 静态数据：品牌 × 系列（来自文档第5页）
# ──────────────────────────────────────────────
BRAND_SERIES: dict[str, list[str]] = {
    "Proya 珀莱雅": [
        "红宝石系列 Ruby",
        "粉宝石系列 Ruby Pink",
        "双抗系列 Dual",
        "启时集致系列 Rich Power",
        "蕴白集光系列 Radiant Whitening",
        "肌源系列 Health Power",
        "光学瓶系列 Professional Whitening",
        "控油抗痘系列 Oil Control",
        "紧致肌密凝时滋养系列 Firming Secret",
        "UV防晒系列",
    ],
    "L'Oreal 欧莱雅": [
        "小蜜罐系列 Honey",
        "复颜玻尿酸水光充盈 Revitalift Filler",
        "黑精华系列",
        "复颜 Revitalift 系列",
        "金致臻颜黑松露 Black Truffle",
        "金致臻颜牡丹 Age Perfect Peony",
        "葡萄籽 Grape Seed & 注白 Whitening",
        "防晒UV系列",
    ],
    "Estee Lauder 雅诗兰黛": [
        "小棕瓶系列 Advanced Night Repair",
        "智妍系列 Revitalizing Supreme+",
        "白金菁萃 Re-Nutriv Ultimate Lift",
        "白金黑钻 Re-Nutriv Ultimate Diamond",
        "专研 Perfectionist Pro & 微精露 Micro Essence & 红石榴 Nutritious",
    ],
    "Kiehl's 科颜氏": [
        "金盏花 Calendula",
        "高保湿 Super Hydration",
        "安白瓶 Brightening & Whitening",
        "紫玻a Retexturizing & Standalone",
    ],
    "SKIN CEUTICALS 修丽可": [
        "发光瓶 Discoloration Defense",
        "植萃色修修护系列",
        "AA系列",
    ],
    "Guyu 谷雨": [
        "光感系列 Whitening Line",
        "雪肌系列",
        "白千松露控油",
        "雪绒修护",
        "山参抗老",
        "月见仙人掌",
        "Post-CP & Cleanser Standalone",
        "Natural Line",
    ],
    "OSM 欧诗漫": [
        "光耀钻白",
        "舒颜修白",
        "肌活复源 & 控油透亮",
        "净透润白",
        "营养透白",
        "金致焕妍",
        "紧致奢颜",
        "水活智润",
        "肌活修护",
        "沁润舒活",
        "面膜系列",
    ],
    "Lancome 兰蔻": [
        "黑金系列",
        "小黑瓶和粉水系列",
        "菁纯系列",
        "极光系列",
        "塑颜系列",
    ],
}

# 护肤步骤（列头）
STEPS: list[tuple[str, str]] = [
    ("Cleanser", "洁面"),
    ("Toner", "水"),
    ("Spray", "喷雾"),
    ("Emulsion", "乳液/精华乳/乳"),
    ("Essence", "精华/精华液"),
    ("Oil Essence", "精华油/油珠精华"),
    ("SUD Essence", "次抛精华/精华棒"),
    ("Cream", "霜"),
    ("Eye", "眼霜"),
    ("Eye Essence", "眼精华"),
    ("Eye Mask", "眼面膜"),
    ("Piece Mask/Plaster", "面膜/敷贴"),
    ("Jar/Gel Mask", "泥膜/面膜（罐装）"),
    ("UV", "防晒"),
    ("Other", "其他"),
]

BRAND_COLORS: dict[str, str] = {
    "Proya 珀莱雅":           "#FF6B6B",
    "L'Oreal 欧莱雅":         "#4ECDC4",
    "Estee Lauder 雅诗兰黛":  "#45B7D1",
    "Kiehl's 科颜氏":         "#96CEB4",
    "SKIN CEUTICALS 修丽可":  "#6C5CE7",
    "Guyu 谷雨":              "#FDCB6E",
    "OSM 欧诗漫":             "#E17055",
    "Lancome 兰蔻":           "#C44569",
}

# ──────────────────────────────────────────────
# 占位数据：Hero SKU 卡片内容
# （实际使用时从 Excel 或爬虫结果加载）
# ──────────────────────────────────────────────
# 格式: HERO_DATA[(brand, series, step_en)] = {"name": str, "price": str, "img_path": str}
HERO_DATA: dict[tuple[str, str, str], dict] = {
    # Proya 红宝石 示例
    ("Proya 珀莱雅", "红宝石系列 Ruby", "Toner"):
        {"name": "珀莱雅红宝石3.0精华水", "price": "¥299", "img_path": ""},
    ("Proya 珀莱雅", "红宝石系列 Ruby", "Essence"):
        {"name": "珀莱雅红宝石3.0精华液", "price": "¥349", "img_path": ""},
    ("Proya 珀莱雅", "红宝石系列 Ruby", "Cream"):
        {"name": "珀莱雅红宝石3.0面霜", "price": "¥359", "img_path": ""},
    ("Proya 珀莱雅", "红宝石系列 Ruby", "Eye"):
        {"name": "珀莱雅红宝石3.0眼霜", "price": "¥259", "img_path": ""},
    # Proya 双抗
    ("Proya 珀莱雅", "双抗系列 Dual", "Toner"):
        {"name": "珀莱雅双抗精华水", "price": "¥199", "img_path": ""},
    ("Proya 珀莱雅", "双抗系列 Dual", "Essence"):
        {"name": "珀莱雅双抗精华液1.0/2.0", "price": "¥259", "img_path": ""},
    ("Proya 珀莱雅", "双抗系列 Dual", "Cream"):
        {"name": "珀莱雅双抗面霜", "price": "¥259", "img_path": ""},
    # SKIN CEUTICALS AA 系列
    ("SKIN CEUTICALS 修丽可", "AA系列", "Essence"):
        {"name": "修丽可维生素CE精华", "price": "¥1620", "img_path": ""},
    ("SKIN CEUTICALS 修丽可", "AA系列", "Toner"):
        {"name": "修丽可RBE夜间精华", "price": "¥1620", "img_path": ""},
    # L'Oreal 小蜜罐
    ("L'Oreal 欧莱雅", "小蜜罐系列 Honey", "Cream"):
        {"name": "欧莱雅金致臻颜花蜜胶原滋润霜", "price": "¥399", "img_path": ""},
}

# ──────────────────────────────────────────────
# 尝试从 Hero_CI_Demo.xlsx 加载真实采集数据
# ──────────────────────────────────────────────
EXCEL_HERO = Path(r"C:\Users\xie.x.3\Documents\Olay CI\Hero_CI_Demo.xlsx")

# 步骤关键词映射（中文产品名 → 护肤步骤）
_STEP_KEYWORDS: list[tuple[str, str]] = [
    ("洁面", "Cleanser"), ("清洁", "Cleanser"), ("卸妆", "Cleanser"),
    ("水", "Toner"), ("爽肤水", "Toner"), ("精华水", "Toner"), ("化妆水", "Toner"),
    ("喷雾", "Spray"),
    ("乳液", "Emulsion"), ("精华乳", "Emulsion"),
    ("次抛", "SUD Essence"), ("精华棒", "SUD Essence"),
    ("精华油", "Oil Essence"), ("油珠", "Oil Essence"),
    ("精华液", "Essence"), ("精华", "Essence"),
    ("眼面膜", "Eye Mask"), ("眼膜", "Eye Mask"),
    ("眼精华", "Eye Essence"),
    ("眼霜", "Eye"), ("眼部", "Eye"),
    ("泥膜", "Jar/Gel Mask"), ("啫喱膜", "Jar/Gel Mask"),
    ("面膜", "Piece Mask/Plaster"), ("贴片", "Piece Mask/Plaster"),
    ("防晒", "UV"), ("UV", "UV"),
    ("面霜", "Cream"), ("凝时", "Cream"), ("霜", "Cream"),
]

_BRAND_MAP_XLSX = {
    "SKIN CEUTICALS": "SKIN CEUTICALS 修丽可",
    "PROYA": "Proya 珀莱雅",
    "ESTEE LAUDER": "Estee Lauder 雅诗兰黛",
    "LOREAL": "L'Oreal 欧莱雅",
    "KIEHLS": "Kiehl's 科颜氏",
    "OSM": "OSM 欧诗漫",
    "GUYU": "Guyu 谷雨",
    "LANCOME": "Lancome 兰蔻",
}

def _guess_step(product_name: str) -> str:
    name = product_name or ""
    for kw, step in _STEP_KEYWORDS:
        if kw in name:
            return step
    return "Other"

@st.cache_data(show_spinner=False)
def load_hero_data() -> dict[tuple[str, str, str], dict]:
    """从 Hero_CI_Demo.xlsx 加载，与静态 HERO_DATA 合并"""
    data = dict(HERO_DATA)  # 从静态数据开始
    if not EXCEL_HERO.exists():
        return data
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_HERO, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            brand_ui = _BRAND_MAP_XLSX.get(sheet_name)
            if not brand_ui:
                continue
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            def col(row, name, fallback=""):
                try:
                    idx = headers.index(name)
                    v = row[idx].value
                    return str(v).strip() if v else fallback
                except Exception:
                    return fallback

            for row in ws.iter_rows(min_row=2):
                pname = col(row, "产品名称")
                if not pname:
                    continue
                price = col(row, "当前价格")
                step = _guess_step(pname)
                # 用第一个 series 作为默认（实际可细化）
                series_list = BRAND_SERIES.get(brand_ui, ["其他"])
                # 简单用第一个 series 填入（Hero SKU 跨品牌矩阵主要关注步骤维度）
                key = (brand_ui, series_list[0], step)
                if key not in data:
                    data[key] = {"name": pname, "price": price or "", "img_path": ""}
    except Exception:
        pass
    return data

# ──────────────────────────────────────────────
# CSS 样式
# ──────────────────────────────────────────────
st.markdown("""
<style>
/* 矩阵表格 */
.hero-table { border-collapse: collapse; width: 100%; font-size: 12px; }
.hero-table th {
    background: #2c3e50; color: #ecf0f1;
    padding: 6px 4px; text-align: center;
    border: 1px solid #34495e; white-space: nowrap; font-size: 11px;
}
.hero-table td {
    border: 1px solid #dfe6e9; padding: 4px;
    vertical-align: top; min-width: 90px; max-width: 130px;
}
.hero-table tr:hover td { background: #f8f9fa; }
.row-header {
    background: #f1f2f6; font-weight: 700;
    writing-mode: horizontal-tb; font-size: 11px;
    padding: 6px 4px; min-width: 130px; max-width: 150px;
}
/* SKU 卡片 */
.sku-card {
    background: #ffffff; border-radius: 6px;
    border-left: 3px solid #6c5ce7;
    padding: 4px 6px; margin: 2px 0;
    font-size: 11px; box-shadow: 0 1px 3px rgba(0,0,0,0.1);
}
.sku-name { font-weight: 600; color: #2d3436; line-height: 1.3; }
.sku-price { color: #e17055; font-weight: 700; margin-top: 2px; }
.empty-cell { color: #b2bec3; text-align: center; font-size: 10px; }
/* 品牌标题 */
.brand-badge {
    display: inline-block; padding: 2px 8px;
    border-radius: 12px; color: white; font-weight: 700;
    font-size: 10px; margin-right: 4px;
}
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────
# 侧边栏：筛选器
# ──────────────────────────────────────────────
st.sidebar.title("🏆 Hero SKU Matrix")
st.sidebar.markdown("---")

all_brands = list(BRAND_SERIES.keys())
selected_brands = st.sidebar.multiselect(
    "选择品牌 Brand",
    options=all_brands,
    default=all_brands[:3],
    help="可多选，默认显示前3个品牌"
)

if not selected_brands:
    st.sidebar.warning("请至少选择一个品牌")
    st.stop()

# 根据选中品牌汇总所有可选系列
all_series_options: list[str] = []
for brand in selected_brands:
    for s in BRAND_SERIES[brand]:
        all_series_options.append(f"{brand} | {s}")

selected_series_keys = st.sidebar.multiselect(
    "选择系列 Series（可选，留空=显示所有）",
    options=all_series_options,
    default=[],
    help="格式：品牌 | 系列名。留空则显示所选品牌的全部系列"
)

# 步骤筛选（默认全选）
step_labels = [f"{en} {zh}" for en, zh in STEPS]
selected_step_labels = st.sidebar.multiselect(
    "选择护肤步骤 Skincare Steps",
    options=step_labels,
    default=step_labels,
)
selected_steps = [STEPS[step_labels.index(lbl)] for lbl in selected_step_labels]

st.sidebar.markdown("---")
st.sidebar.caption("数据来源：天猫旗舰店爬虫 + Skincare Competitor Product Menu Mar'26")

# ──────────────────────────────────────────────
# 主体：矩阵表格
# ──────────────────────────────────────────────
st.title("🏆 Hero SKU Cross Brand Matrix")
st.caption("列：护肤步骤   行：品牌 × 系列   格：明星单品卡片")

if not selected_steps:
    st.warning("请在左侧选择至少一个护肤步骤")
    st.stop()

hero_data = load_hero_data()

# 构建要展示的 (brand, series) 行列表
rows: list[tuple[str, str]] = []
if selected_series_keys:
    for key in selected_series_keys:
        brand, series = key.split(" | ", 1)
        rows.append((brand, series))
else:
    for brand in selected_brands:
        for series in BRAND_SERIES[brand]:
            rows.append((brand, series))

# ── 渲染 HTML 表格 ─────────────────────────────
def _card(brand: str, series: str, step_en: str, color: str) -> str:
    info = hero_data.get((brand, series, step_en))
    if not info:
        return '<span class="empty-cell">—</span>'
    name = info.get("name", "")[:30]
    price = info.get("price", "")
    price_str = f'<div class="sku-price">{price}</div>' if price else ""
    return (
        f'<div class="sku-card" style="border-left-color:{color}">'
        f'<div class="sku-name">{name}</div>'
        f'{price_str}'
        f'</div>'
    )

# 构建列头
th_rows = ""
for en, zh in selected_steps:
    th_rows += f'<th>{en}<br><span style="font-weight:300;font-size:10px">{zh}</span></th>'

# 构建数据行
tbody = ""
prev_brand = None
for brand, series in rows:
    color = BRAND_COLORS.get(brand, "#636e72")
    # 品牌分组标题行
    if brand != prev_brand:
        colspan = len(selected_steps) + 1
        badge = f'<span class="brand-badge" style="background:{color}">{brand}</span>'
        tbody += (
            f'<tr><td colspan="{colspan}" style="background:#f0f0f0;'
            f'padding:4px 8px;font-weight:700;">'
            f'{badge}</td></tr>'
        )
        prev_brand = brand

    # 数据行
    cells = f'<td class="row-header">{series}</td>'
    for en, _ in selected_steps:
        cells += f'<td>{_card(brand, series, en, color)}</td>'
    tbody += f'<tr>{cells}</tr>'

table_html = f"""
<div style="overflow-x:auto">
<table class="hero-table">
<thead>
  <tr>
    <th style="min-width:140px">品牌 / 系列</th>
    {th_rows}
  </tr>
</thead>
<tbody>
{tbody}
</tbody>
</table>
</div>
"""

st.markdown(table_html, unsafe_allow_html=True)

# ── 数据统计 ──────────────────────────────────
total_cells = len(rows) * len(selected_steps)
filled = sum(
    1 for brand, series in rows
    for en, _ in selected_steps
    if hero_data.get((brand, series, en))
)
st.markdown("---")
col1, col2, col3 = st.columns(3)
col1.metric("品牌数", len(selected_brands))
col2.metric("系列数", len(rows))
col3.metric(f"已填充 SKU 卡片", f"{filled} / {total_cells}",
            delta=f"{filled/total_cells*100:.0f}%" if total_cells else "0%")

st.info(
    "💡 **如何填充 SKU 卡片**：运行 `python -m scraper.main` 采集天猫旗舰店数据后，"
    "重新刷新本页即可自动读取 `Hero_CI_Demo.xlsx` 中的采集结果。"
    "当前卡片为静态示例数据。"
)
