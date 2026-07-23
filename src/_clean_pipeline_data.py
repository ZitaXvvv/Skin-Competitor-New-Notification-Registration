"""
清理 CI_List_Ada.xlsx 中的垃圾数据：
1. 删除产品名只有品牌名、含"--"、或为"-"的行
2. 删除 2026-05/06/07 的 pipeline 旧数据（保留今天 07/23/2026 导入的行）
"""
import openpyxl
import re
import shutil
from pathlib import Path
from datetime import date, datetime

DOCS = Path(r"C:\Users\xie.x.3\Documents\Olay CI")
MAIN = DOCS / "CI_List_Ada.xlsx"

# 已知品牌名（用于检测"只有品牌名"的行）
BRAND_ONLY = {
    "珀莱雅", "谷雨", "欧诗漫", "科颜氏", "欧莱雅", "雅诗兰黛",
    "修丽可", "兰蔻", "百雀羚", "自然堂", "薇诺娜", "韩束", "娇韵诗",
    "妮维雅", "资生堂", "Nivea", "Shiseido",
}

TODAY = "07/23/2026"  # 今天导入的行的 upload_time

def is_bad_name(name_raw) -> bool:
    if not name_raw:
        return True
    name = str(name_raw).strip()
    if not name or name == "-":
        return True
    if "--" in name:          # 品牌-- 分类 格式
        return True
    # 名字就是品牌名本身（4字以内 + 完全匹配已知品牌）
    if name in BRAND_ONLY:
        return True
    return False

def is_2026_mayjunjul_pipeline(row) -> bool:
    """判断是否是 2026年5/6/7月 的旧 pipeline 数据。
    只看 notif_raw（D列，产品真实的备案/通知时间），不再看 upload_raw（A列，
    只是抓取脚本运行的日期，不代表产品本身的日期——之前用它来判断删除，
    曾导致某一天抓取的全部合法数据被误删）。
    """
    upload_raw = row[0]  # A 列：upload time（仅用于"今天导入"的保留豁免，不用于删除判断）
    notif_raw  = row[3]  # D 列：notification time（真实备案/通知日期）

    # 如果今天导入的（upload = 07/23/2026），保留
    if upload_raw and str(upload_raw).strip() == TODAY:
        return False

    # 只检查备案/通知时间是否在 2026-05/06/07
    val = notif_raw
    if val is None:
        return False
    if hasattr(val, "month"):
        return val.year == 2026 and val.month in (5, 6, 7)
    s = str(val).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            d = datetime.strptime(s, fmt)
            return d.year == 2026 and d.month in (5, 6, 7)
        except ValueError:
            pass
    return False

# 删除前先备份原文件，防止清理规则出错导致数据不可逆丢失
backup_path = DOCS / f"CI_List_Ada_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
shutil.copy2(MAIN, backup_path)
print(f"[BACKUP] 已备份原文件到 {backup_path.name}")

wb = openpyxl.load_workbook(MAIN)
total_deleted = 0

for sh in wb.sheetnames:
    ws = wb[sh]
    rows_to_keep = [list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]]  # 表头
    deleted = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(c is None for c in row):
            continue
        name = row[1] if len(row) > 1 else None
        
        if is_bad_name(name):
            deleted += 1
            continue
        if is_2026_mayjunjul_pipeline(row):
            deleted += 1
            continue
        rows_to_keep.append(row)
    
    if deleted > 0:
        # 重写 sheet
        ws.delete_rows(1, ws.max_row)
        for i, r in enumerate(rows_to_keep, 1):
            for j, v in enumerate(r, 1):
                ws.cell(i, j, v)
        print(f"  [{sh}] 删除 {deleted} 行，保留 {len(rows_to_keep)-1} 行")
        total_deleted += deleted
    else:
        print(f"  [{sh}] 无需清理（{ws.max_row-1} 行）")

wb.save(MAIN)
wb.close()
print(f"\n[DONE] 共删除 {total_deleted} 行，已保存 {MAIN.name}")
