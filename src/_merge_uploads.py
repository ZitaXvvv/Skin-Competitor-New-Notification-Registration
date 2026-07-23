"""
_merge_uploads.py
-----------------
把用户上传的各品牌 Excel（5列格式）及 CI_List_Ada.xlsx（10列格式）
合并进主文件 C:\\Users\\xie.x.3\\Documents\\Olay CI\\CI_List_Ada.xlsx

规则：
  - 不覆盖已有行
  - 按备案号（reg_num）去重
  - 样稿链接 → col J（mini POC）；若是 NMPA PDF 则同时写 col H（link）
"""

import shutil
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# ── 路径配置 ─────────────────────────────────────────────
DEST   = Path(r"C:\Users\xie.x.3\Documents\Olay CI\CI_List_Ada.xlsx")
TEMP   = Path(r"C:\Users\xie.x.3\AppData\Local\Temp\1")

# 品牌文件名 → 目标 sheet 名
BRAND_FILES = {
    "雅诗兰黛.xlsx":  "ESTEE LAUDER",
    "薇诺娜.xlsx":    "Winona",
    "珀莱雅.xlsx":    "PROYA",
    "欧莱雅.xlsx":    "LOREAL",
    "科颜氏.xlsx":    "Kiehls",
    "娇韵诗.xlsx":    "Clains",
}

# CI_List_Ada.xlsx 里的 sheet → 目标 sheet 的映射（名字相同）
CILIST_FILE = "CI_List_Ada.xlsx"

# 标准 10 列表头（与主文件一致）
STD_HEADERS = (
    "upload time", "Name", "English / Benefit", "Notification Time",
    "#", "Registration Time", "Ingredient", "link",
    "化妆品产品标签样稿", "mini POC",
)

TODAY = date.today().strftime("%m/%d/%Y")

# ── 工具函数 ──────────────────────────────────────────────

import re as _re
# 备案号识别：覆盖国产/进口/特殊各类格式
_REG_PAT = _re.compile(
    r"妆网备字|妆网备进字|国妆网备进字|国妆备字|国妆备进字|"
    r"国妆特字|国妆特进字|卫妆特字|省妆备字",
    _re.IGNORECASE,
)

def norm_reg(v) -> str:
    """标准化备案号：去空格、全角→半角"""
    s = str(v or "").strip()
    return s.replace("（", "(").replace("）", ")").replace("\xa0", " ").strip()


def is_pdf_url(url: str) -> bool:
    """判断是否是 NMPA 直链 PDF"""
    return "nmpa.gov.cn/datasearch" in url or url.endswith(".pdf")


def fmt_date(v) -> str:
    """统一日期为 MM/DD/YYYY 字符串"""
    from datetime import datetime as _dt
    if v is None:
        return ""
    if isinstance(v, (_dt,)):
        return v.strftime("%m/%d/%Y")
    try:
        from datetime import date as _d
        if isinstance(v, _d):
            return v.strftime("%m/%d/%Y")
    except Exception:
        pass
    return str(v).strip()


def ensure_sheet(wb: openpyxl.Workbook, sheet_name: str):
    """确保 sheet 存在，不存在则新建并写表头"""
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(list(STD_HEADERS))
        print(f"  [新建] sheet: {sheet_name}")
    return wb[sheet_name]


def load_existing_regs(ws) -> set:
    """从已有 sheet 收集所有备案号（col E = index 4），用于去重"""
    regs = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 5:
            for cell in row:
                s = norm_reg(cell)
                if s and _REG_PAT.search(s):
                    regs.add(s)
    return regs


# ── 合并品牌文件（5列格式）───────────────────────────────

def merge_brand_file(src: Path, sheet_name: str, dest_wb: openpyxl.Workbook):
    """从 5列品牌文件 读取并追加到目标 sheet"""
    try:
        src_wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ❌ 无法打开 {src.name}: {e}")
        return 0

    ws_dest = ensure_sheet(dest_wb, sheet_name)
    existing = load_existing_regs(ws_dest)

    added = skipped = 0
    for sh in src_wb.sheetnames:
        ws_src = src_wb[sh]
        for row in ws_src.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            name      = str(row[0] or "").strip()
            art_url   = str(row[1] or "").strip()
            category  = str(row[2] or "").strip()
            reg_num   = norm_reg(row[3])
            notif_raw = row[4] if len(row) > 4 else None

            if not name or not reg_num:
                continue
            if reg_num in existing:
                skipped += 1
                continue

            notif_str = fmt_date(notif_raw)

            # 决定 link（col H）和 mini POC（col J）
            link_h = ""
            poc_j  = ""
            if art_url:
                if is_pdf_url(art_url):
                    link_h = art_url   # 特殊化妆品 PDF → Artwork
                else:
                    poc_j = art_url    # 普通备案产品详情页 → POC

            new_row = [
                TODAY,       # A: upload time
                name,        # B: Name
                "",          # C: English / Benefit
                notif_str,   # D: Notification Time
                reg_num,     # E: #
                category,    # F: Registration Time (类目)
                "",          # G: Ingredient
                link_h,      # H: link
                "",          # I: 化妆品产品标签样稿
                poc_j,       # J: mini POC
            ]
            ws_dest.append(new_row)
            existing.add(reg_num)
            added += 1

    src_wb.close()
    print(f"  {src.name} → {sheet_name}: +{added} 新增, {skipped} 已跳过(重复)")
    return added


# ── 合并 CI_List_Ada.xlsx（10列格式）──────────────────────

def merge_cilist_file(src: Path, dest_wb: openpyxl.Workbook):
    """从 10列 CI_List_Ada 格式文件 合并所有 sheet"""
    try:
        src_wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ❌ 无法打开 {src.name}: {e}")
        return

    total_added = 0
    for sheet_name in src_wb.sheetnames:
        ws_src = src_wb[sheet_name]
        ws_dest = ensure_sheet(dest_wb, sheet_name)
        existing = load_existing_regs(ws_dest)

        added = skipped = 0
        for row in ws_src.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5:
                continue
            if not any(c is not None for c in row):
                continue

            # 找备案号（扫全行）
            reg_num = ""
            for cell in row:
                s = norm_reg(cell)
                if s and _REG_PAT.search(s):
                    reg_num = s
                    break

            if not reg_num:
                continue
            if reg_num in existing:
                skipped += 1
                continue

            # 标准化日期列（col A 和 col D）
            def g(i):
                return row[i] if i < len(row) else None

            new_row = [
                fmt_date(g(0)) or TODAY,  # A
                str(g(1) or "").strip(),  # B name
                str(g(2) or "").strip(),  # C benefit
                fmt_date(g(3)),           # D notification date
                reg_num,                  # E reg_num
                str(g(5) or "").strip(),  # F registration time / category
                str(g(6) or "").strip(),  # G ingredient
                str(g(7) or "").strip(),  # H link
                str(g(8) or "").strip(),  # I label
                str(g(9) or "").strip(),  # J poc
            ]
            ws_dest.append(new_row)
            existing.add(reg_num)
            added += 1
            total_added += 1

        if added or skipped:
            print(f"  [{sheet_name}]: +{added} 新增, {skipped} 已跳过")

    src_wb.close()
    return total_added


# ── 主流程 ────────────────────────────────────────────────

def main():
    print(f"目标文件: {DEST}")

    # 0. 写入前自动备份，避免合并出错导致数据不可恢复
    backup_dir = DEST.parent / "_admin_backups"
    backup_dir.mkdir(exist_ok=True)
    backup_path = backup_dir / f"{DEST.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy2(DEST, backup_path)
    print(f"已备份到: {backup_path}")

    dest_wb = openpyxl.load_workbook(str(DEST))

    # 1. 合并品牌文件
    print("\n=== 合并品牌文件 ===")
    for fname, sheet_name in BRAND_FILES.items():
        src = TEMP / fname
        if not src.exists():
            print(f"  ⚠️  {fname} 不存在，跳过")
            continue
        merge_brand_file(src, sheet_name, dest_wb)

    # 2. 合并 CI_List_Ada.xlsx（Jul'26 版本）
    ci_src = TEMP / CILIST_FILE
    if ci_src.exists():
        print(f"\n=== 合并 {CILIST_FILE} ===")
        merge_cilist_file(ci_src, dest_wb)
    else:
        print(f"\n⚠️  {CILIST_FILE} 不存在，跳过")

    # 3. 保存
    dest_wb.save(str(DEST))
    print(f"\n✅ 已保存: {DEST}")


if __name__ == "__main__":
    main()
