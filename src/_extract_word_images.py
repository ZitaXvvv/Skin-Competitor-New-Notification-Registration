"""
_extract_word_images.py
从两份竞品产品菜单 Word 文档中提取图片，
匹配 Excel 里当前无图的产品，写入 image_map.json。

用法:
  python src/_extract_word_images.py [--dry-run]

逻辑:
  1. 解析两份 docx 的列式表格 (每列=一个SKU，name行+image行)
  2. 对每个 Excel 品牌，找出尚无图片的产品
  3. 对每条无图产品，用去品牌前缀的子串匹配 Word 里提取到的名称
  4. 匹配成功 → 保存为 JPEG ≤320px → 写入 image_map.json
  5. 去重：已有图的产品不覆盖
"""

import argparse, json, re, zipfile, io
import xml.etree.ElementTree as ET
from pathlib import Path
from PIL import Image

# ── 路径配置 ──────────────────────────────────────────────────────────────
WORKSPACE = Path(__file__).parent.parent
DOCS = [
    Path(r"C:\Users\xie.x.3\OneDrive - Procter and Gamble\CBE 2025") / "Skincare Competitor Product Menu Mar'26.docx",
    Path(r"C:\Users\xie.x.3\OneDrive - Procter and Gamble\CBE 2025") / "Skincare Competitor Product Menu Nov'25.docx",
]
MAP_JSON  = WORKSPACE / "res" / "product_images" / "image_map.json"
OUT_DIR   = WORKSPACE / "res" / "product_images" / "_word"
EXCEL_PATH = Path(r"C:\Users\xie.x.3\Documents\Olay CI\CI_List_Ada.xlsx")

OUT_DIR.mkdir(parents=True, exist_ok=True)

MAX_DIM_PX = 320
JPEG_Q     = 82

# XML 命名空间
NS_W  = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
NS_A  = "http://schemas.openxmlformats.org/drawingml/2006/main"
NS_R  = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
R_EMBED = f"{{{NS_R}}}embed"

BRAND_MAP = {
    "珀莱雅": "PROYA",   "谷雨": "GUYU",       "欧诗漫": "OSM",
    "兰蔻":   "Lancome", "欧莱雅": "LOREAL",   "雅诗兰黛": "ESTEE LAUDER",
    "修丽可": "SKIN CEUTICALS", "百雀羚": "BQL",
    "韩束":   "Kans",    "自然堂": "Chando",    "薇诺娜": "Winona",
    "科颜氏": "Kiehls",  "契尔氏": "Kiehls",    "娇韵诗": "Clains",
}
# longest-prefix-first for brand stripping
_BRAND_LIST = sorted(BRAND_MAP.keys(), key=len, reverse=True)

NAME_LABELS  = ("产品中文名", "中文名", "产品名称", "ChineseName", "ChinaName",
                "产品名", "China Name")
IMG_LABELS_PRI = ("正装图片", "RegularSizePic", "正装照片", "产品图片")
IMG_LABELS_SEC = ("小样照片", "系列图", "MiniPic", "迷你照片")
BAD_NAME_RE    = re.compile(r"^(Size|是否有样品|N/?A|\d+m?[Ll].*)$", re.IGNORECASE)


def norm(s: str) -> str:
    return re.sub(r"\s+", "", s.strip())


def cell_text(el) -> str:
    return "".join(t.text or "" for t in el.findall(f".//{{{NS_W}}}t")).strip()


def cell_rids(el) -> list:
    return [b.get(R_EMBED) for b in el.findall(f".//{{{NS_A}}}blip") if b.get(R_EMBED)]


def detect_brand(text: str) -> str:
    for zh in _BRAND_LIST:
        if zh in text:
            return BRAND_MAP[zh]
    return ""


def strip_brand(name: str) -> str:
    for zh in _BRAND_LIST:
        if name.startswith(zh):
            return name[len(zh):]
    return name


def safe_fname(s: str) -> str:
    return re.sub(r'[\\/:*?"<>|\x00-\x1f]', "_", s.strip())[:100]


def save_jpg(raw_bytes: bytes, out_path: Path) -> bool:
    try:
        img = Image.open(io.BytesIO(raw_bytes)).convert("RGB")
        w, h = img.size
        if max(w, h) > MAX_DIM_PX:
            r = MAX_DIM_PX / max(w, h)
            img = img.resize((max(1, int(w * r)), max(1, int(h * r))), Image.LANCZOS)
        img.save(str(out_path), "JPEG", quality=JPEG_Q, optimize=True)
        return True
    except Exception as e:
        print(f"    保存失败 {out_path.name}: {e}")
        return False


# ── 解析单份 docx，返回 {brand_en: {prod_name_in_doc: raw_bytes}} ─────────
def parse_docx(docx_path: Path) -> dict:
    print(f"\n解析: {docx_path.name}  ({docx_path.stat().st_size // 1024 // 1024} MB)")
    result: dict = {}   # brand_en -> {name -> bytes}

    with zipfile.ZipFile(str(docx_path)) as zf:
        doc_xml  = zf.read("word/document.xml").decode("utf-8", errors="replace")
        rels_xml = zf.read("word/_rels/document.xml.rels").decode("utf-8", errors="replace")

        rid_to_media: dict = {}
        for rel in ET.fromstring(rels_xml):
            tgt = rel.get("Target", "")
            if "media" in tgt:
                rid_to_media[rel.get("Id")] = "word/" + tgt.replace("../", "")

        root   = ET.fromstring(doc_xml)
        tables = root.findall(f".//{{{NS_W}}}tbl")
        print(f"  共 {len(tables)} 个表格")

        matched_total = 0
        for tbl in tables:
            rows = tbl.findall(f".//{{{NS_W}}}tr")
            if len(rows) < 2:
                continue

            # 构建 {norm_label: [(text, rids), ...per-column]}
            row_by_label: dict = {}
            for row in rows:
                cells = row.findall(f"./{{{NS_W}}}tc")
                parsed = [(cell_text(c), cell_rids(c)) for c in cells]
                if parsed:
                    lbl = norm(parsed[0][0])
                    if lbl and lbl not in row_by_label:
                        row_by_label[lbl] = parsed

            # Find name row
            name_row = None
            for lbl in NAME_LABELS:
                key = norm(lbl)
                if key in row_by_label:
                    name_row = row_by_label[key]
                    break
            if name_row is None:
                continue

            # Find image row
            img_row = None
            for lbl in IMG_LABELS_PRI + IMG_LABELS_SEC:
                key = norm(lbl)
                if key in row_by_label:
                    img_row = row_by_label[key]
                    break
            if img_row is None:
                continue

            # Detect table brand
            full_text = " ".join(ct for cells in row_by_label.values() for ct, _ in cells)
            table_brand = detect_brand(full_text)

            for ci in range(1, len(name_row)):
                prod_name = name_row[ci][0].strip()
                if not prod_name or len(prod_name) < 3 or BAD_NAME_RE.match(prod_name):
                    continue
                rids = img_row[ci][1] if ci < len(img_row) else []
                if not rids:
                    continue
                brand_en = detect_brand(prod_name) or table_brand
                if not brand_en:
                    continue
                media = rid_to_media.get(rids[0], "")
                if not media or media not in zf.namelist():
                    continue
                try:
                    raw = zf.read(media)
                except Exception:
                    continue
                result.setdefault(brand_en, {})[prod_name] = raw
                matched_total += 1

    print(f"  提取到 {matched_total} 个产品图")
    return result


# ── 产品名匹配 ───────────────────────────────────────────────────────────
def fuzzy_match(excel_name: str, word_names: list) -> str | None:
    """
    尝试找 word_names 中与 excel_name 最接近的名称。
    策略: exact > 去品牌后exact > 子串(≥4字) > None
    """
    clean_excel = strip_brand(excel_name)

    # exact
    if excel_name in word_names:
        return excel_name
    if clean_excel in word_names:
        return clean_excel

    # substring (word_name ∈ excel_name  OR  excel_name ∈ word_name)
    best = None
    best_len = 0
    for wn in word_names:
        clean_wn = strip_brand(wn)
        cand = ""
        if len(clean_wn) >= 4 and clean_wn in clean_excel:
            cand = clean_wn
        elif len(clean_excel) >= 4 and clean_excel in clean_wn:
            cand = clean_excel
        if cand and len(cand) > best_len:
            best, best_len = wn, len(cand)

    return best


# ── 加载 Excel 中无图产品 ────────────────────────────────────────────────
def _all_excel_files() -> list:
    """返回所有 CI_List_Ada*.xlsx（历史 + 当前），与 dashboard.py 保持一致"""
    base = EXCEL_PATH.parent
    historical = sorted(base.glob("CI_List_Ada *.xlsx"))
    current = EXCEL_PATH
    return [f for f in historical if f != current] + ([current] if current.exists() else [])


def load_missing(img_map: dict) -> dict:
    """返回 {brand_en: [prod_name]} 只含当前无图的产品（覆盖全部历史文件）"""
    import openpyxl
    import sys; sys.path.insert(0, str(WORKSPACE / "src"))
    from config import BRANDS

    missing: dict = {}
    seen: set = set()   # 跨文件去重（与 dashboard.py 逻辑一致）

    for xlsx in _all_excel_files():
        try:
            wb = openpyxl.load_workbook(str(xlsx), read_only=True)
        except Exception:
            continue
        for brand_en in BRANDS:
            if brand_en not in wb.sheetnames:
                continue
            ws = wb[brand_en]
            brand_imgs = img_map.get(brand_en, {})
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[1]:
                    continue
                name = str(row[1]).strip()
                key = f"{brand_en}:{name}"
                if key in seen:
                    continue
                seen.add(key)
                if _has_img(name, brand_imgs):
                    continue
                missing.setdefault(brand_en, []).append(name)
        wb.close()

    total = sum(len(v) for v in missing.values())
    print(f"\n共 {total} 个产品当前无图:")
    for b, names in sorted(missing.items()):
        print(f"  {b}: {len(names)}")
    return missing


def _has_img(name: str, brand_imgs: dict) -> bool:
    if name in brand_imgs:
        return True
    clean = strip_brand(name)
    for k in brand_imgs:
        ck = strip_brand(k)
        if len(ck) >= 4 and (ck in clean or clean in ck):
            return True
        if len(clean) >= 4 and clean in k:
            return True
    return False


# ── 主函数 ────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    img_map: dict = json.loads(MAP_JSON.read_text(encoding="utf-8")) if MAP_JSON.exists() else {}
    missing = load_missing(img_map)

    # 解析两份 Word 文档 → 合并
    word_data: dict = {}   # brand_en -> {name -> bytes}
    for doc in DOCS:
        if not doc.exists():
            print(f"[跳过] 文件不存在: {doc}")
            continue
        parsed = parse_docx(doc)
        for brand_en, items in parsed.items():
            word_data.setdefault(brand_en, {})
            for name, raw in items.items():
                # 不覆盖已有 (Mar'26 优先，先解析的优先)
                word_data[brand_en].setdefault(name, raw)

    print(f"\n两份文档合计提取品牌: {list(word_data.keys())}")

    # 匹配
    added = 0
    no_match = []

    for brand_en, excel_names in missing.items():
        word_items = word_data.get(brand_en, {})
        word_name_list = list(word_items.keys())
        if not word_name_list:
            no_match.extend([(brand_en, n) for n in excel_names])
            continue

        for excel_name in excel_names:
            wn = fuzzy_match(excel_name, word_name_list)
            if wn is None:
                no_match.append((brand_en, excel_name))
                continue

            raw = word_items[wn]
            if args.dry_run:
                print(f"  [DRY] [{brand_en}] '{excel_name}'  ←  '{wn}'")
                added += 1
                continue

            # save image
            out_path = OUT_DIR / f"{safe_fname(excel_name)}.jpg"
            if save_jpg(raw, out_path):
                img_map.setdefault(brand_en, {})[excel_name] = str(out_path).replace("\\", "/")
                print(f"  ✓ [{brand_en}] {excel_name}")
                added += 1
            else:
                no_match.append((brand_en, excel_name))

    if not args.dry_run and added:
        MAP_JSON.write_text(json.dumps(img_map, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"\n✅ image_map.json 已更新，新增 {added} 张图片")

    print(f"\n汇总: 新增={added}, 未匹配={len(no_match)}")
    if no_match:
        print("\n未匹配产品 (Word文档中无对应图片):")
        for brand_en, name in no_match[:30]:
            print(f"  [{brand_en}] {name}")
        if len(no_match) > 30:
            print(f"  ... 及其余 {len(no_match)-30} 条")


if __name__ == "__main__":
    main()
