"""
模块1：抓取美丽修行大数据 (bebd.bevol.com)
- 搜索各品牌最新备案/注册产品（护肤 + 防晒）
- 提取产品名、备案/注册号、功效、日期
- 特殊化妆品 → 查询 nmpa.gov.cn/datasearch → 获取 PDF URL
- 普通化妆品 → PDF URL 由模块2补全
- 写入 Excel（去重，不覆盖已有记录）

【首次运行】：脚本会打开可见浏览器，等待你手动登录 bebd.bevol.com
【后续运行】：自动读取保存的 Cookie，无需手动登录
"""

import json
import logging
import random
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import openpyxl
from playwright.sync_api import Page, BrowserContext, sync_playwright

sys.path.insert(0, str(Path(__file__).parent))
from config import (
    BEBD_URL,
    BRANDS,
    COL_CATEGORY,
    COL_DATE,
    COL_EFFECT,
    COL_INGREDIENTS,
    COL_LABEL_URL,
    COL_NAME,
    COL_PDF_URL,
    COL_POC_URL,
    COL_REG_NUM,
    COL_UPLOAD_DATE,
    COOKIES_FILE,
    DOWNLOAD_BASE,
    EXCEL_PATH,
    NMPA_DATASEARCH_URL,
    SEARCH_CATEGORIES,
    TIME_PERIOD_DAYS,
)

log = logging.getLogger(__name__)

EXCEL_HEADERS = [
    "upload time",          # A: 脚本运行当天
    "Name",                 # B: 产品名称
    "English / Benefit",   # C: 功效宣称
    "Notification Time",   # D: 备案时间
    "#",                   # E: 类目（护肤/防晒）
    "Registration Ti",     # F: 备案/注册号（去重依据）
    "Ingredient",          # G: 成分列表（模块2补全）
    "link",                # H: 备案/注册链接 ← 流程块2 value[7]
    "化妆品产品标签链接",    # I: Artwork / 产品标签 PDF
    "mini POC",            # J: 功效证明链接
]


# ─────────────────────────────────────────────
# Cookie 管理
# ─────────────────────────────────────────────

def load_cookies(context: BrowserContext) -> bool:
    if not COOKIES_FILE.exists():
        return False
    with open(COOKIES_FILE, "r", encoding="utf-8") as f:
        cookies = json.load(f)
    context.add_cookies(cookies)
    log.info("已加载保存的 Cookie")
    return True


def save_cookies(context: BrowserContext):
    cookies = context.cookies()
    with open(COOKIES_FILE, "w", encoding="utf-8") as f:
        json.dump(cookies, f, ensure_ascii=False, indent=2)
    log.info(f"Cookie 已保存 → {COOKIES_FILE}")


def is_logged_in(page: Page) -> bool:
    """简单检测是否已登录（不同时期页面结构可能变化，可调整选择器）"""
    try:
        # 登录后通常有"退出"、"我的账户"等元素
        page.wait_for_selector(
            "a:has-text('退出'), .user-avatar, .user-info, span:has-text('登出')",
            timeout=4000,
        )
        return True
    except Exception:
        return False


def ensure_login(page: Page, context: BrowserContext, interactive: bool = True):
    """
    如果未登录：
    - interactive=True（默认，白天手动运行）：等待用户在浏览器里完成登录
    - interactive=False（无人值守计划任务）：抛出 RuntimeError，由调用方跳过
    """
    if is_logged_in(page):
        return
    if not interactive:
        raise RuntimeError(
            "BEBD Cookie 已失效且当前为无人值守模式。"
            "请在工作时段运行 scripts\\refresh_bebd_login.ps1 刷新登录状态。"
        )
    print("\n" + "=" * 60)
    print("【需要手动操作】：")
    print("  浏览器已打开 https://bebd.bevol.com/")
    print("  请在浏览器窗口中完成登录（账号密码或扫码）。")
    print("  登录完成后，回到此终端，按 Enter 继续。")
    print("=" * 60)
    input("登录完成后按 Enter ▶ ")
    save_cookies(context)
    log.info("Cookie 已保存，下次运行将自动登录")


# ─────────────────────────────────────────────
# 备案/注册号分类（决定去哪个NMPA网站）
# ─────────────────────────────────────────────

def classify_reg_num(reg_num: str) -> dict:
    """
    根据备案/注册号中的关键字判断类型，决定查询哪个网站哪个页面。

    国产普通化妆品备案 → hzpba.nmpa.gov.cn 国产页  (含"备"，不含"进")
    进口普通化妆品备案 → hzpba.nmpa.gov.cn 进口页  (含"备"，含"进")
    国产特殊化妆品注册 → nmpa.gov.cn/datasearch    (含"特"，不含"进")
    进口特殊化妆品注册 → nmpa.gov.cn/datasearch    (含"特"，含"进")

    示例：
      浙G妆网备字2026008093  → hzpba 国产
      国妆备进字2026000XXX   → hzpba 进口
      国妆特字20263800       → nmpa  国产特殊
      国妆特进字2026000XXX   → nmpa  进口特殊
    """
    is_imported = "进" in reg_num
    if "特" in reg_num:
        site = "nmpa_datasearch"
        label = ("进口" if is_imported else "国产") + "特殊化妆品注册"
        nmpa_radio = "进口特殊化妆品注册信息" if is_imported else "国产特殊化妆品注册信息"
    elif "备" in reg_num:
        site = "hzpba"
        label = ("进口" if is_imported else "国产") + "普通化妆品备案"
        nmpa_radio = None
    else:
        site = "unknown"
        label = "未知类型"
        nmpa_radio = None
    return {
        "site":        site,
        "label":       label,
        "is_imported": is_imported,
        "nmpa_radio":  nmpa_radio,   # NMPA datasearch 里要选的 radio 文字
    }


# ─────────────────────────────────────────────
# 美丽修行大数据 爬取
# ─────────────────────────────────────────────

def parse_date(text: str) -> Optional[date]:
    for fmt in ("%Y.%m.%d", "%Y/%m/%d", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text.strip(), fmt).date()
        except ValueError:
            pass
    log.warning(f"无法解析日期: {text!r}")
    return None


def _drag_slider(page: Page, popup_locator, drag_pixels: float,
                  slider_btn_sel: str = "") -> bool:
    """执行人性化滑块拖动，成功返回 True。"""
    btn_sels = [
        slider_btn_sel,
        ".yidun_slider__icon", "i.yidun_slider__icon",
        "div[class*='slider__icon']", ".yidun_slider",
        "div[class*='yidun_slider']",
    ]
    btn_box = None
    for sel in btn_sels:
        if not sel:
            continue
        el = page.locator(sel).first
        if el.count() > 0:
            btn_box = el.bounding_box()
            if btn_box:
                break

    if not btn_box:
        log.warning("  找不到滑块按钮元素")
        return False

    x0 = btn_box["x"] + btn_box["width"] / 2
    y0 = btn_box["y"] + btn_box["height"] / 2

    # ease-in-out 曲线 + 随机抖动，模拟人手
    page.mouse.move(x0, y0)
    page.wait_for_timeout(random.randint(200, 400))
    page.mouse.down()
    page.wait_for_timeout(random.randint(100, 200))

    steps = 35
    for i in range(steps + 1):
        t = i / steps
        ease = t * t * (3 - 2 * t)   # cubic ease-in-out
        page.mouse.move(
            x0 + drag_pixels * ease + random.uniform(-0.5, 0.5),
            y0 + random.uniform(-1.5, 1.5),
        )
        page.wait_for_timeout(random.randint(5, 18))

    page.wait_for_timeout(random.randint(300, 500))
    page.mouse.up()
    page.wait_for_timeout(2000)
    return not popup_locator.is_visible()


def _handle_yidun_captcha(page: Page) -> bool:
    """
    检测并处理易盾验证码弹窗。
    策略（按优先级）：
      A. 点关闭蒙层直接跳过
      B. ddddocr（本地免费）：截取背景图+滑块图 → 识别偏移 → 自动拖动
      C. 超级鹰（在线服务）：截整体图 → 识别偏移 → 自动拖动
      D. 手动：提示用户 90 秒内在浏览器里操作
    返回 True 表示验证码已消失（或从未出现）。
    """
    popup = page.locator(".yidun_popup, .yidun_modal, .yidun")
    if popup.count() == 0:
        return True
    try:
        popup.first.wait_for(state="visible", timeout=1500)
    except Exception:
        return True

    log.warning("  ⚠️ 检测到易盾验证弹窗")
    page.bring_to_front()
    page.wait_for_timeout(1000)   # 等动画完成

    # ── 方案A：点关闭蒙层 ──
    for close_sel in [
        ".yidun_popup__mask[aria-label='关闭']",
        ".yidun_modal__mask[aria-label='关闭']",
        ".yidun_close",
    ]:
        try:
            btn = page.locator(close_sel).first
            if btn.count() > 0:
                btn.click(force=True)
                page.wait_for_timeout(1000)
                if not popup.first.is_visible():
                    log.info("  ✅ 弹窗已关闭（点蒙层）")
                    return True
        except Exception:
            pass

    cap_dir = Path(__file__).parent.parent / "log"
    cap_dir.mkdir(parents=True, exist_ok=True)

    # ── 方案B：ddddocr 本地识别 ──
    try:
        import ddddocr  # noqa: PLC0415

        # 找背景图元素（带缺口的大图）
        bg_el = None
        for sel in [".yidun_bgimg", "img[class*='bgimg']",
                    ".yidun_slide-bg", ".yidun_bg"]:
            el = page.locator(sel).first
            if el.count() > 0:
                bg_el = el
                break

        # 找滑块图元素（要嵌入缺口的小图）
        piece_el = None
        for sel in [".yidun_jigsaw-img", "img[class*='jigsaw']",
                    ".yidun_slide-piece", ".yidun_block img",
                    ".yidun_slider img"]:
            el = page.locator(sel).first
            if el.count() > 0:
                piece_el = el
                break

        if bg_el and piece_el:
            bg_bytes    = bg_el.screenshot()
            piece_bytes = piece_el.screenshot()

            slide_det = ddddocr.DdddOcr(det=False, ocr=False, show_ad=False)
            result = slide_det.slide_match(piece_bytes, bg_bytes)
            target_x_img = result["target"][0]  # 图片坐标系中的目标 x
            log.info(f"  ddddocr 目标 x: {target_x_img}px（图片坐标）")

            # 换算到浏览器坐标中的拖动距离
            bg_box    = bg_el.bounding_box() or {"width": 300}
            natural_w = bg_el.evaluate("el => el.naturalWidth") or bg_box["width"]
            scale = bg_box["width"] / natural_w
            drag_px = target_x_img * scale
            log.info(f"  拖动距离: {drag_px:.1f}px（浏览器坐标）")

            if _drag_slider(page, popup.first, drag_px):
                log.info("  ✅ 验证码已通过（ddddocr）")
                return True
            log.warning("  ddddocr 未通过，尝试超级鹰…")
        else:
            log.warning(f"  ddddocr: 找不到图片元素 bg={bg_el is not None} piece={piece_el is not None}")

    except ImportError:
        log.info("  ddddocr 未安装，跳过")
    except Exception as e:
        log.warning(f"  ddddocr 方案失败: {e}")

    # ── 方案C：超级鹰在线识别 ──
    try:
        cap_path = str(cap_dir / "yidun_captcha.png")
        captcha_area = page.locator(".yidun_modal, .yidun_popup, .yidun").first
        captcha_area.screenshot(path=cap_path)
        log.info(f"  验证码截图: {cap_path}")

        sys.path.insert(0, str(Path(__file__).parent.parent / "extend" / "python"))
        from pass_captcha import slider as chaojiying_slider
        offset = chaojiying_slider(cap_path)
        log.info(f"  超级鹰识别偏移: {offset}px")

        if offset > 0 and _drag_slider(page, popup.first, float(offset)):
            log.info("  ✅ 验证码已通过（超级鹰）")
            return True
        log.warning("  超级鹰也未通过，转为手动")
    except Exception as e:
        log.warning(f"  超级鹰方案失败: {e}")

    # ── 方案D：手动（最长等 90 秒）──
    log.warning("  ⚠️  请在浏览器中手动完成验证码（最长等 90 秒）…")
    for _ in range(90):
        time.sleep(1)
        try:
            if not popup.first.is_visible():
                log.info("  ✅ 验证码已完成（手动）")
                return True
        except Exception:
            return True
    log.error("  ❌ 验证码超时，跳过")
    return False



def _apply_filter(page: Page, category: str) -> bool:
    """点开商品分类级联选择器，选择指定类目，点筛选"""
    try:
        # 先处理可能存在的易盾验证码弹窗
        _handle_yidun_captcha(page)

        # 点开级联选择器（商品分类 dropdown）
        cascader = page.locator("input.ant-cascader-input").first
        if cascader.count() == 0:
            log.warning(f"  未找到商品分类筛选器")
            return False
        cascader.click(timeout=8000)
        page.wait_for_timeout(800)

        # 在弹出的下拉菜单里点对应类目
        option = page.locator(
            f".ant-cascader-menu-item:has-text('{category}'), "
            f"li.ant-cascader-menu-item:has-text('{category}')"
        ).first
        if option.count() == 0:
            log.warning(f"  级联选择器中未找到：{category}")
            page.keyboard.press("Escape")
            return False
        option.click()
        page.wait_for_timeout(500)

        # 点"筛选"按钮
        page.locator("button:has-text('筛选')").first.click(timeout=8000)
        page.wait_for_timeout(2500)
        log.info(f"  ✅ 筛选器已选择：{category}")
        return True
    except Exception as exc:
        log.error(f"  应用筛选器({category})出错: {exc}")
        return False


def _reset_filter(page: Page):
    """点"重置"按钮清除筛选器"""
    try:
        reset = page.locator("button:has-text('重置')").first
        if reset.count() > 0:
            reset.click()
            page.wait_for_timeout(1500)
    except Exception:
        pass


def _sort_by_date(page: Page):
    """
    点击页面顶部"备案时间"排序 Tab（综合排序|安全|用户评分|备案时间|美修指数）。
    策略1: 用 JS 找到元素坐标 → 用 page.mouse.click 真实鼠标点击（对 Vue 组件最可靠）。
    策略2: 降级用 Playwright locator 的 .click()。
    """
    try:
        _handle_yidun_captcha(page)

        # 等待结果行出现（确保 Tab 已渲染）
        try:
            page.wait_for_selector(".ant-table-row", timeout=8000)
        except Exception:
            pass

        # 额外等待 Vue 渲染排序 Tab
        page.wait_for_timeout(800)

        # ── 策略1：用 JS 获取坐标，再用真实鼠标点击 ──
        rect_info = page.evaluate("""
            () => {
                // 查找所有含"备案时间"文本的叶子级元素（排除表格内）
                const all = [...document.querySelectorAll('span, li, div, a, label, button')];
                const matches = [];
                for (const el of all) {
                    // 跳过含子元素太多的容器
                    if (el.children.length > 2) continue;
                    const txt = el.textContent.trim();
                    if ((txt === '备案时间' || txt.startsWith('备案时间'))
                            && !el.closest('table')
                            && !el.closest('.ant-table')
                            && !el.closest('th')) {
                        const r = el.getBoundingClientRect();
                        matches.push({
                            tag: el.tagName,
                            cls: el.className,
                            txt: txt.slice(0, 30),
                            x: r.x + r.width / 2,
                            y: r.y + r.height / 2,
                            visible: r.width > 0 && r.height > 0,
                        });
                    }
                }
                return matches;
            }
        """)
        log.info(f"  备案时间 Tab 候选: {rect_info}")

        clicked = False
        if rect_info:
            for item in rect_info:
                if item.get("visible") and item.get("x", 0) > 0 and item.get("y", 0) > 0:
                    cls = item.get("cls", "")
                    # 已经是 active（降序）则无需再点，再点会切回升序
                    if "active" in cls:
                        log.info(f"  备案时间 Tab 已激活，跳过点击 → {item['tag']}.{cls[:40]}")
                        clicked = True
                        break
                    x, y = item["x"], item["y"]
                    log.info(f"  鼠标点击 ({x:.0f}, {y:.0f}) → {item['tag']}.{cls[:40]}")
                    page.mouse.click(x, y)
                    page.wait_for_timeout(1500)
                    clicked = True
                    break

        # ── 策略2：用 Playwright locator 降级 ──
        if not clicked:
            log.warning("  策略1未命中，尝试 Playwright locator")
            try:
                # get_by_text 自动跳过不可见元素
                tabs = page.get_by_text("备案时间", exact=True).all()
                log.info(f"  Playwright get_by_text 找到 {len(tabs)} 个")
                for tab in tabs:
                    try:
                        # 排除 table 内的同名元素
                        in_table = tab.evaluate(
                            "el => !!el.closest('table, .ant-table, th')"
                        )
                        if not in_table:
                            tab.click(timeout=4000)
                            page.wait_for_timeout(1500)
                            clicked = True
                            log.info("  Playwright locator 点击成功")
                            break
                    except Exception as e:
                        log.debug(f"  locator click 失败: {e}")
            except Exception as e:
                log.warning(f"  策略2也失败: {e}")

        if not clicked:
            log.warning("  ⚠️ 未能点击备案时间排序 Tab，结果可能未按日期排序")

    except Exception as exc:
        log.warning(f"  排序出错: {exc}")


def _extract_rows(page: Page, brand_en: str, cutoff: date, today: date) -> list[dict]:
    """提取当前页所有符合日期窗口的数据行"""
    try:
        page.wait_for_selector(".ant-table-row", timeout=8000)
    except Exception:
        log.info("    无数据行（超时）")
        return []

    rows = page.locator(".ant-table-row").all()
    log.info(f"    当前页: {len(rows)} 行")

    # 调试：打印前3行原始数据（含各列内容，用于确认列索引）
    for _i, _row in enumerate(rows[:3]):
        _c = _row.locator("td").all()
        if len(_c) >= 10:
            _cell_texts = [_c[j].inner_text().strip()[:30] for j in range(min(5, len(_c)))]
            log.info(f"      [样本{_i}] cells[0..4]={_cell_texts}  备案={_c[8].inner_text().strip()!r}  日期={_c[9].inner_text().strip()!r}")

    results = []
    for row in rows:
        cells = row.locator("td").all()
        if len(cells) < 10:
            log.debug(f"    跳过(列数不足{len(cells)})")
            continue
        try:
            date_text = cells[9].inner_text().strip()
            prod_date = parse_date(date_text)
            if not prod_date:
                log.debug(f"    跳过(日期解析失败: {date_text!r})")
                continue
            if prod_date > today:
                log.debug(f"    跳过(日期{prod_date}在未来)")
                continue
            if prod_date < cutoff:
                log.debug(f"    跳过(日期{prod_date}早于截止{cutoff})")
                continue

            reg_num = cells[8].inner_text().strip()
            if not reg_num:
                log.debug(f"    跳过(备案号为空)")
                continue
            if "出" in reg_num:
                log.debug(f"    跳过(备案号含'出': {reg_num!r})")
                continue

            # 产品名：
            # cells[1]/cells[2] 内可能有 <a> 链接包含完整"品牌--产品名"格式
            name = ""
            for _ci in [1, 2]:
                if _ci < len(cells):
                    _link = cells[_ci].locator("a").first
                    if _link.count() > 0:
                        _lt = _link.inner_text().strip()
                        if _lt and len(_lt) > 3:
                            name = _lt
                            break
            if not name:
                # 降级：品牌名 + "--" + 类目（可作为唯一标识）
                _brand = cells[2].inner_text().strip()
                _cat   = cells[5].inner_text().strip()
                name = f"{_brand}--{_cat}" if _brand and _cat else (_brand or reg_num)
            if not name:
                log.debug(f"    跳过(产品名为空)")
                continue
            # 搜索科颜氏时混入契尔氏、搜兰蔻时混入名女人，均为串台结果直接跳过
            if "契尔氏" in name or "名女人" in name:
                log.debug(f"    跳过(非目标品牌串台: {name!r})")
                continue

            results.append({
                "date":        prod_date.strftime("%m/%d/%Y"),
                "name":        name,
                "reg_num":     reg_num,
                "effect":      cells[7].inner_text().strip(),
                "category":    cells[5].inner_text().strip(),
                "brand_en":    brand_en,
                "is_special":  "特" in reg_num,
                "is_imported": "进" in reg_num,
                "pdf_url":     None,
            })
        except Exception as exc:
            log.warning(f"    解析行出错: {exc}")
    log.info(f"    本页符合条件: {len(results)} 条（截止日期{cutoff}，今天{today}）")
    return results


def search_brand(page: Page, brand_cn: str, brand_en: str) -> list[dict]:
    """
    完整搜索流程（每个品牌）：
      1. 搜索品牌名
      2. 双击"备案时间"排序（降序，最新在前）——只做一次
      3. 对每个类目（防晒、护肤）：
         a. 应用筛选器（继承已有排序）
         b. 翻页抓取
         c. 重置筛选器（无论是否出错都执行）
    """
    today = date.today()
    cutoff = today - timedelta(days=TIME_PERIOD_DAYS)
    all_results: list[dict] = []
    seen_regs: set[str] = set()

    try:
        # ── 1. 搜索品牌 ──
        # 先处理搜索页面可能已弹出的验证码
        _handle_yidun_captcha(page)

        # 验证码完成后等待页面稳定（防止 Vue 重置搜索框内容）
        page.wait_for_timeout(800)

        # 重试最多3次填入品牌名（fill 后检查值是否真的进去了）
        for _attempt in range(3):
            search_box = page.locator("input[placeholder='查找化妆品']").first
            search_box.click(timeout=8000)
            page.wait_for_timeout(300)
            # 先清空，再逐字输入（比 fill 更可靠）
            search_box.click(click_count=3)  # 全选
            page.keyboard.press("Delete")
            search_box.type(brand_cn, delay=80)
            page.wait_for_timeout(400)
            current_val = search_box.input_value()
            log.info(f"  搜索框输入 (尝试{_attempt+1}): {current_val!r}")
            if current_val.strip() == brand_cn.strip():
                break
            log.warning(f"  输入未成功，重试…")
            page.wait_for_timeout(600)
        else:
            log.error(f"  [{brand_en}] 无法输入品牌名，跳过")
            return []

        # 再次检测（填字后可能触发验证）
        _handle_yidun_captcha(page)
        page.wait_for_timeout(400)

        # 点击搜索按钮（兜底：button 找不到时用 Enter 键）
        try:
            btn = page.locator("button.search-btn").first
            btn.wait_for(state="visible", timeout=4000)
            btn.click(timeout=6000)
        except Exception:
            log.warning("  button.search-btn 不可点，改用 Enter 键搜索")
            try:
                search_box.press("Enter")
            except Exception:
                # 最后兜底：重新导航到搜索页再按 Enter
                log.warning("  重新导航到搜索页再重试")
                _navigate_to_search_page(page)
                page.wait_for_timeout(1000)
                sb2 = page.locator("input[placeholder='查找化妆品']").first
                sb2.click(click_count=3)
                page.keyboard.press("Delete")
                sb2.type(brand_cn, delay=80)
                sb2.press("Enter")
        page.wait_for_timeout(3000)

        # 搜索结果加载后可能再触发验证
        _handle_yidun_captcha(page)
        page.wait_for_timeout(500)

        # 排序在 filter 后执行，此处不做（表格还未加载，列头不可见）

    except Exception as exc:
        log.error(f"  [{brand_en}] 搜索/排序出错: {exc}", exc_info=True)
        return []

    # ── 3. 对每个类目逐一筛选 ──
    for category in SEARCH_CATEGORIES:   # ["防晒", "护肤"]
        log.info(f"\n  [{brand_en}] 类目: {category}")
        try:
            if not _apply_filter(page, category):
                continue

            # filter 应用后表格已加载，现在排序（列头已可见）
            _sort_by_date(page)
            for page_num in range(5):
                log.info(f"    第 {page_num+1} 页")
                page_rows = _extract_rows(page, brand_en, cutoff, today)

                # 若最新行已超出时间窗口，后续页也不会有新数据
                rows_all = page.locator(".ant-table-row").all()
                if rows_all:
                    _c = rows_all[0].locator("td").all()
                    if len(_c) >= 10:
                        first_date = parse_date(_c[9].inner_text().strip())
                        if first_date and first_date < cutoff:
                            log.info(f"    最新日期({first_date})超出窗口，停止翻页")
                            for prod in page_rows:
                                if prod["reg_num"] not in seen_regs:
                                    seen_regs.add(prod["reg_num"])
                                    all_results.append(prod)
                            break

                for prod in page_rows:
                    if prod["reg_num"] not in seen_regs:
                        seen_regs.add(prod["reg_num"])
                        all_results.append(prod)

                # 翻页
                next_li = page.locator("li.ant-pagination-next").first
                if next_li.count() == 0:
                    break
                cls  = next_li.get_attribute("class") or ""
                aria = next_li.get_attribute("aria-disabled") or ""
                if "disabled" in cls or aria == "true":
                    break
                try:
                    next_li.locator("button").click(timeout=5000)
                    page.wait_for_timeout(2000)
                except Exception:
                    log.info("    翻页失败，停止")
                    break

        except Exception as exc:
            log.error(f"  [{brand_en}][{category}] 出错: {exc}", exc_info=True)
        finally:
            # 无论成功或出错，都重置筛选器，保证下一个类目从干净状态开始
            _reset_filter(page)
            page.wait_for_timeout(1000)

    log.info(f"  {brand_en}: 共找到 {len(all_results)} 条新品（过去 {TIME_PERIOD_DAYS} 天）")
    return all_results


# ─────────────────────────────────────────────
# NMPA 国家局 特殊化妆品查询
# ─────────────────────────────────────────────

def get_nmpa_special_pdf_url(page: Page, reg_num: str, is_imported: bool) -> Optional[str]:
    """
    在 nmpa.gov.cn/datasearch 查询特殊化妆品注册信息（用备案/注册号查询）。
    进入"当前详情"页，返回 Artwork PDF 预览 URL。
    注意：此函数使用调用方传入的 page 对象（已在 run() 中忽略 SSL 错误）。
    """
    try:
        # 设置更长超时重试导航（NMPA 页面加载较慢，重试3次）
        for _attempt in range(3):
            try:
                page.goto(NMPA_DATASEARCH_URL, timeout=60000)
                page.wait_for_timeout(5000)
                # 如果页面有内容了就继续
                if page.locator("body").inner_text().strip():
                    break
                log.debug(f"  NMPA 页面尚空，等待后重试 ({_attempt+1}/3)")
                page.wait_for_timeout(5000)
            except Exception:
                page.wait_for_timeout(4000)

        # 诊断：若页面仍为空则截图并退出
        body_text = page.locator("body").inner_text().strip()
        if not body_text or len(body_text) < 20:
            log.warning(f"  NMPA 页面内容为空，可能加载失败")
            try:
                diag = str(Path(__file__).parent.parent / "log" / "nmpa_diag.png")
                page.screenshot(path=diag, full_page=True)
                log.info(f"  已保存诊断截图: {diag}")
            except Exception:
                pass
            return None

        # 关闭引导弹窗
        for sel in ["a:has-text('关闭')", ".close-guide", ".modal-close", "button:has-text('关闭')"]:
            close = page.locator(sel).first
            if close.count() > 0:
                try:
                    close.click(timeout=2000)
                    page.wait_for_timeout(500)
                    break
                except Exception:
                    pass

        # 点击化妆品入口（兼容 a/div/li/span 等多种元素类型）
        cos_sel = (
            "a:has-text('化妆品'), li:has-text('化妆品'), "
            "div.category-item:has-text('化妆品'), span.nav-item:has-text('化妆品'), "
            "[class*='category']:has-text('化妆品')"
        )
        try:
            page.wait_for_selector(cos_sel, timeout=20000)
            page.locator(cos_sel).first.click()
            page.wait_for_timeout(2000)
        except Exception as e:
            log.warning(f"  未找到化妆品入口: {e}")
            try:
                diag_path = str(Path(__file__).parent.parent / "log" / "nmpa_diag.png")
                page.screenshot(path=diag_path)
                log.info(f"  NMPA页面截图已保存: {diag_path}")
            except Exception:
                pass
            return None

        # 选择注册类型 — 多策略点击（元素可能不可见，需 scroll + force）
        reg_type = "进口特殊化妆品注册信息" if is_imported else "国产特殊化妆品注册信息"
        label_clicked = False
        # 先等一等让页面渲染完
        page.wait_for_timeout(1000)
        # 策略1：scroll into view 后点 <label>
        label = page.locator(f"label:has-text('{reg_type}')").first
        if label.count() > 0:
            try:
                label.scroll_into_view_if_needed(timeout=5000)
                page.wait_for_timeout(500)
                label.click(force=True, timeout=8000)
                label_clicked = True
                log.info(f"  ✅ 已选择注册类型: {reg_type}")
            except Exception as e:
                log.debug(f"  策略1失败: {e}")
        if not label_clicked:
            # 策略2：向上找 label 祖先
            try:
                span = page.locator(f"text={reg_type}").first
                parent_label = span.locator("xpath=ancestor::label[1]")
                if parent_label.count() > 0:
                    parent_label.scroll_into_view_if_needed(timeout=3000)
                    parent_label.click(force=True, timeout=5000)
                    label_clicked = True
                    log.info(f"  ✅ 已选择注册类型(策略2): {reg_type}")
            except Exception as e:
                log.debug(f"  策略2失败: {e}")
        if not label_clicked:
            # 策略3：直接 force click 文本元素
            try:
                el = page.locator(f"text={reg_type}").first
                el.scroll_into_view_if_needed(timeout=3000)
                el.click(force=True, timeout=5000)
                label_clicked = True
                log.info(f"  ✅ 已选择注册类型(策略3): {reg_type}")
            except Exception as e:
                log.warning(f"  无法选择注册类型 {reg_type}: {e}")
                return None
        page.wait_for_timeout(1500)

        # 用备案/注册号查询（比产品名精准）
        search_box = page.locator(
            "input#productName, input#registrationNum, "
            "input[placeholder*='注册证'], input[placeholder*='注册号'], "
            "input[placeholder*='产品名称'], input[placeholder*='请输入']"
        ).first
        if search_box.count() == 0:
            search_box = page.locator("input[type='text']").nth(0)
        search_box.fill(reg_num)
        page.locator(
            "button:has-text('查询'), button:has-text('搜索'), "
            "input[type='submit'], .search-btn"
        ).first.click()
        page.wait_for_timeout(8000)

        # 检查是否有结果
        no_data = page.locator(
            "text=未查到相关信息, text=暂无数据, .no-data, "
            "td:has-text('未查询到'), td:has-text('没有数据')"
        ).first
        if no_data.count() > 0:
            log.info(f"  国家局无注册记录: {reg_num}")
            return None

        # 点击"当前详情"
        detail_link = page.locator(
            "text=当前详情, span:has-text('当前详情'), a:has-text('详情'), "
            "td:has-text('当前详情') a, .detail-btn"
        ).first
        if detail_link.count() == 0:
            log.warning(f"  未找到'当前详情'按钮: {reg_num}")
            return None
        detail_link.click()
        page.wait_for_timeout(6000)

        # 在详情页查找 PDF 链接
        for link in page.locator("a[href*='preview-pdf'], a[href*='.pdf']").all():
            href = link.get_attribute("href") or ""
            if href:
                if href.startswith("/"):
                    href = "https://www.nmpa.gov.cn" + href
                log.info(f"  找到特殊化妆品PDF: {href[:80]}")
                return href

        for el in page.locator("embed[src], iframe[src]").all():
            src = el.get_attribute("src") or ""
            if ".pdf" in src or "preview" in src:
                return src

        current_url = page.url
        if "preview-pdf" in current_url or ".pdf" in current_url:
            return current_url

        log.warning(f"  详情页未找到PDF链接: {reg_num}")
        return None

    except Exception as exc:
        log.error(f"国家局查询出错（{reg_num}）: {exc}", exc_info=True)
        return None


# ─────────────────────────────────────────────
# Excel 写入
# ─────────────────────────────────────────────

def ensure_sheet(wb: openpyxl.Workbook, sheet_name: str):
    """确保 sheet 存在且有表头"""
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        for col, header in enumerate(EXCEL_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
    return wb[sheet_name]


def write_products_to_excel(all_products: dict):
    """将各品牌新产品写入 Excel（去重）"""
    Path(EXCEL_PATH).parent.mkdir(parents=True, exist_ok=True)

    if Path(EXCEL_PATH).exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
    else:
        log.warning(f"Excel 文件不存在，将新建: {EXCEL_PATH}")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # 删除默认空 sheet

    for brand_en, products in all_products.items():
        ws = ensure_sheet(wb, brand_en)

        # 读取已有备案号（去重用）
        # 搜索整行任意列，兼容不同历史版本写入位置不一致的情况
        import re as _re
        _REG_PAT = _re.compile(r"(妆网备字|国妆备字|国妆特字|国妆特进字|国妆备进字|卫妆特字)")
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            for cell in row:
                s = str(cell or "").strip()
                if s and _REG_PAT.search(s):
                    existing.add(s)
                    break  # 每行只取第一个匹配的备案号

        written = 0
        today_str = date.today().strftime("%m/%d/%Y")
        for prod in products:
            if prod["reg_num"] in existing:
                log.debug(f"  跳过重复: {prod['reg_num']}")
                continue

            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=COL_UPLOAD_DATE, value=today_str)       # A
            ws.cell(row=next_row, column=COL_NAME,        value=prod["name"])    # B
            ws.cell(row=next_row, column=COL_EFFECT,      value=prod["effect"])  # C
            ws.cell(row=next_row, column=COL_DATE,        value=prod["date"])    # D 备案时间
            ws.cell(row=next_row, column=COL_CATEGORY,    value=prod["category"]) # E
            ws.cell(row=next_row, column=COL_REG_NUM,     value=prod["reg_num"]) # F
            # G (Ingredient) 空白，由模块2从 NMPA 补全
            if prod.get("pdf_url"):
                ws.cell(row=next_row, column=COL_PDF_URL, value=prod["pdf_url"]) # H link
            # I 化妆品产品标签链接 / J mini POC 空白，由后续模块娹

            existing.add(prod["reg_num"])
            written += 1

        log.info(f"  {brand_en}: 写入 {written} 条新记录（总共 {ws.max_row - 1} 条）")

    # 确保所有品牌都有 sheet
    for brand_en in BRANDS.keys():
        ensure_sheet(wb, brand_en)

    wb.save(EXCEL_PATH)
    log.info(f"Excel 已保存 → {EXCEL_PATH}")


# ─────────────────────────────────────────────
# 搜索页健康检查 & 恢复
# ─────────────────────────────────────────────

def _navigate_to_search_page(page: Page):
    """确保 bebd_page 处于搜索页（有 '查找化妆品' 输入框）"""
    # 如果已经在搜索页，直接返回
    try:
        if page.locator("input[placeholder='查找化妆品']").first.count() > 0:
            return
    except Exception:
        pass
    # 重新导航到 BEBD 首页，然后点搜索菜单
    try:
        log.info("  [页面恢复] 重新导航到 BEBD 搜索页…")
        current_url = page.url
        if not current_url.startswith("https://bebd.bevol.com"):
            page.goto(BEBD_URL, timeout=30000)
            page.wait_for_timeout(2000)
        page.locator("li.ant-menu-item:has(.menu-search)").first.click(timeout=8000)
        page.wait_for_selector("input[placeholder='查找化妆品']", timeout=10000)
        log.info("  [页面恢复] 搜索页已恢复")
    except Exception as e:
        log.warning(f"  [页面恢复] 失败: {e}")


# ─────────────────────────────────────────────
# 主入口
# ─────────────────────────────────────────────

def run(headless_override: Optional[bool] = None, unattended: bool = False):
    """
    运行模块1。

    headless_override=None → 有 Cookie 文件则 headless，否则显示浏览器等待登录
    headless_override=True → 强制 headless（需已有有效 Cookie）
    headless_override=False → 强制显示浏览器
    unattended=True → 无人值守模式（计划任务）：Cookie 失效时跳过 BEBD 而不阻塞等待
    """
    has_cookies = COOKIES_FILE.exists()
    # 默认可见模式，方便手动处理验证码；无人值守时强制 headless
    headless = headless_override if headless_override is not None else (True if unattended else False)
    log.info(f"启动 Playwright (headless={headless}, unattended={unattended})")

    all_products: dict = {brand: [] for brand in BRANDS}

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=headless,
                args=[
                    "--lang=zh-CN",
                    "--disable-blink-features=AutomationControlled",
                    "--no-sandbox",
                ],
            )
            context = browser.new_context(
                locale="zh-CN",
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
            )

            # 登录 bebd
            bebd_page = context.new_page()
            load_cookies(context)
            bebd_page.goto(BEBD_URL, timeout=30000)
            bebd_page.wait_for_timeout(2000)
            ensure_login(bebd_page, context, interactive=not unattended)

            # 主页搜索框是 <div>，不是 <input>，需先点左侧"搜索"菜单进入搜索页
            log.info("导航到搜索页…")
            bebd_page.locator("li.ant-menu-item:has(.menu-search)").first.click()
            bebd_page.wait_for_timeout(2000)
            # 确认搜索页已加载（有 '查找化妆品' 输入框）
            try:
                bebd_page.wait_for_selector("input[placeholder='查找化妆品']", timeout=8000)
                log.info("搜索页已加载")
            except Exception:
                log.warning("未检测到搜索输入框，继续尝试…")

            # 用于查询 NMPA 的独立页面
            # NMPA 页面用独立 context，忽略 SSL 证书错误（公司网络代理）
            nmpa_ctx = browser.new_context(
                locale="zh-CN",
                ignore_https_errors=True,
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
            )
            nmpa_page = nmpa_ctx.new_page()

            for brand_en, brand_cn in BRANDS.items():
                log.info(f"\n{'─'*50}\n处理品牌: {brand_en}（{brand_cn}）")

                # 每个品牌前做健康检查，防止 NMPA 查询后 BEBD 页面状态丢失
                _navigate_to_search_page(bebd_page)

                brand_products = search_brand(bebd_page, brand_cn, brand_en)

                log.info(f"  共找到 {len(brand_products)} 条新品")

                # 根据备案号关键字决定去哪个NMPA网站
                for prod in brand_products:
                    cls = classify_reg_num(prod["reg_num"])
                    prod["_cls"] = cls   # 缓存分类结果

                    if cls["site"] == "nmpa_datasearch":
                        log.info(f"  [{cls['label']}] 查询国家局datasearch: {prod['reg_num']}")
                        url = get_nmpa_special_pdf_url(
                            nmpa_page, prod["reg_num"], cls["is_imported"]
                        )
                        prod["pdf_url"] = url
                        time.sleep(2)
                    elif cls["site"] == "hzpba":
                        log.info(f"  [{cls['label']}] 将由模块2查询 hzpba: {prod['reg_num']}")
                        # 模块2 负责 hzpba 查询并回写 Excel
                    else:
                        log.warning(f"  [未知类型] 无法判断网站: {prod['reg_num']}")

                all_products[brand_en] = brand_products

            nmpa_page.close()
            nmpa_ctx.close()
            bebd_page.close()
            browser.close()

    except RuntimeError as _login_err:
        # 无人值守模式下 Cookie 失效时的优雅退出（不阻塞计划任务）
        log.warning(f"⚠️  BEBD 跳过（无人值守模式）: {_login_err}")
        log.warning("   请在工作时段运行 scripts\\refresh_bebd_login.ps1 刷新登录。")
    except Exception as _run_exc:
        log.error(f"运行中断: {_run_exc}", exc_info=True)
    finally:
        # 无论是否出错，都把已收集的数据写入 Excel
        write_products_to_excel(all_products)
        log.info("\n✅ 模块1 完成（已写入 Excel）")


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )
    run()
