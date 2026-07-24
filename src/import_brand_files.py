"""
一次性导入脚本：将品牌单独 Excel 文件合并到 CI_List_Ada.xlsx
使用方法：python src/import_brand_files.py
"""
import openpyxl
import re
import shutil
from datetime import datetime
from pathlib import Path

TEMP = Path(r"C:\Users\xie.x.3\AppData\Local\Temp\1")
DOCS = Path(r"C:\Users\xie.x.3\Documents\Olay CI")
MAIN = DOCS / "CI_List_Ada.xlsx"

# 品牌文件 → Excel sheet 名称映射
BRAND_MAP = {
    "百雀羚":   "BQL",
    "谷雨":     "GUYU",
    "韩束":     "Kans",
    "兰蔻":     "Lancome",
    "欧诗漫":   "OSM",
    "修丽可":   "SKIN CEUTICALS",
    "自然堂":   "Chando",
    "雅诗兰黛": "ESTEE LAUDER",
    "珀莱雅":   "PROYA",
    "娇韵诗":   "Clains",
    "薇诺娜":   "Winona",
    "科颜氏":   "Kiehls",
}

HEADERS = [
    "upload time", "Name", "English / Benefit", "Notification Time",
    "#", "Registration Time", "Ingredient", "link",
    "化妆品产品标签样稿", "mini POC",
]

REG_PAT = re.compile(
    r"(妆网备字|国妆备字|国妆特字|国妆特进字|国妆备进字|卫妆特字|网备进字|国妆网备进字)"
)

# 与 dashboard.py 保持一致：导入时同步排除男士/唇部/契尔氏/名女人产品
EXCLUDE_PAT = re.compile(r"唇|口红|男士|男仕|契尔氏|名女人")


def clean_reg(s: str) -> str:
    # 统一全角括号 → 半角，再去除空白，确保去重一致
    t = str(s or "").replace("（", "(").replace("）", ")")
    return re.sub(r"\s+", "", t).strip()


def main():
    today_str = datetime.today().strftime("%m/%d/%Y")

    # ── Step 1: 另存 CI_List_Ada Jul'26.xlsx ──
    uploaded = TEMP / "CI_List_Ada.xlsx"
    if uploaded.exists():
        hist = DOCS / "CI_List_Ada Jul'26.xlsx"
        shutil.copy2(uploaded, hist)
        print(f"[OK] 历史档已保存: {hist.name}  ({hist.stat().st_size:,} bytes)")
    else:
        print(f"[WARN] 未找到上传的 CI_List_Ada.xlsx，跳过历史档保存")

    # ── Step 2: 打开主文件 ──
    if not MAIN.exists():
        print(f"[ERROR] 主文件不存在: {MAIN}")
        return
    wb = openpyxl.load_workbook(MAIN)
    print(f"[OK] 主文件已加载: {MAIN.name}  sheets={wb.sheetnames}")

    total_new = 0

    # ── Step 3: 逐品牌导入 ──
    for cn_name, sheet_name in BRAND_MAP.items():
        src = TEMP / f"{cn_name}.xlsx"
        if not src.exists():
            print(f"  [SKIP] 未找到: {src.name}")
            continue

        # 确保 sheet 存在，有表头
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            for col, h in enumerate(HEADERS, 1):
                ws.cell(row=1, column=col, value=h)
            print(f"  [NEW] 创建 sheet: {sheet_name}")
        ws = wb[sheet_name]

        # 收集已有备案号
        existing: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            for cell in row:
                s = clean_reg(cell)
                if s and REG_PAT.search(s):
                    existing.add(s)
                    break

        # 读源文件：A=商品名称, B=样稿链接, C=商品分类, D=备案号, E=备案时间
        src_wb = openpyxl.load_workbook(src, read_only=True)
        src_ws = src_wb.active
        written = 0
        for row in src_ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            name     = str(row[0]).strip()
            link     = str(row[1] or "").strip()
            category = str(row[2] or "").strip()
            reg_raw  = str(row[3] or "").strip()
            date_raw = row[4]

            reg = clean_reg(reg_raw)
            if not reg or not REG_PAT.search(reg):
                continue
            if EXCLUDE_PAT.search(name):          # 排除男士/唇部产品
                continue
            if reg in existing:
                continue

            # 日期
            if hasattr(date_raw, "strftime"):
                date_str = date_raw.strftime("%m/%d/%Y")
            else:
                date_str = str(date_raw or "").strip()

            nr = ws.max_row + 1
            ws.cell(nr, 1, today_str)                                   # A upload time
            ws.cell(nr, 2, name)                                        # B Name
            ws.cell(nr, 4, date_str)                                    # D Notification Time
            ws.cell(nr, 5, reg)                                         # E 备案号
            ws.cell(nr, 6, category)                                    # F 分类
            ws.cell(nr, 8, link if link not in ("None", "NA", "") else "")  # H link

            existing.add(reg)
            written += 1

        src_wb.close()
        print(f"  [{sheet_name}] +{written} 新行 (已有 {len(existing)-written} 条去重)")
        total_new += written

    # ── Step 4: 保存 ──
    wb.save(MAIN)
    wb.close()
    print(f"\n[DONE] 共新增 {total_new} 条记录 → {MAIN}")


if __name__ == "__main__":
    main()
